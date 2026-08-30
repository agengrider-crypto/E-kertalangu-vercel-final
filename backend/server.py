from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import re
import uuid
import base64
import logging
import secrets
import hmac
import hashlib
import time
import math
import shutil
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Literal, Annotated, Any

import bcrypt
import jwt
import qrcode
from openpyxl import Workbook, load_workbook
from fastapi import FastAPI, APIRouter, Depends, HTTPException, Request, status, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict

# --------- Env & Mongo ---------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALG = 'HS256'
TOKEN_TTL_DAYS = int(os.environ.get('TOKEN_TTL_DAYS', '30'))
LOGIN_LOCK_WINDOW_SEC = 15 * 60
LOGIN_MAX_ATTEMPTS = 5
ADMIN_CONTACT_WA = os.environ.get('ADMIN_CONTACT_WA', '6281937718541')

logger = logging.getLogger('ekertalangu')
logging.basicConfig(level=logging.INFO)

# --------- Utils ---------

def now_utc() -> datetime:
    return datetime.now(timezone.utc)


WITA_OFFSET = timedelta(hours=8)


def now_wita() -> datetime:
    """Waktu WITA (UTC+8) naive — server berjalan UTC, absensi & jam laporan pakai WITA."""
    return (datetime.now(timezone.utc) + WITA_OFFSET).replace(tzinfo=None)


def iso(dt: datetime) -> str:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


def new_id() -> str:
    return uuid.uuid4().hex[:20]


def slug(s: str) -> str:
    return re.sub(r'[^a-z0-9]+', '-', (s or '').lower()).strip('-')[:40]


def hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_pw(pw: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode('utf-8'), h.encode('utf-8'))
    except Exception:
        return False


def normalize_phone(p: Optional[str]) -> Optional[str]:
    if not p:
        return None
    p = re.sub(r'\D', '', p)
    if p.startswith('0'):
        p = '62' + p[1:]
    return p


def make_token(user_id: str, role: str, token_version: int = 0, extra: dict = None) -> str:
    payload = {
        'sub': user_id,
        'role': role,
        'tv': int(token_version or 0),
        'exp': now_utc() + timedelta(days=TOKEN_TTL_DAYS),
        'iat': now_utc(),
        'type': 'access',
    }
    if extra:
        payload.update(extra)
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


def qr_png_bytes(text: str) -> bytes:
    img = qrcode.make(text)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return buf.getvalue()


def qr_png_datauri(text: str) -> str:
    return 'data:image/png;base64,' + base64.b64encode(qr_png_bytes(text)).decode('ascii')


# --------- Log helper ---------
async def log_activity(actor_id: str, actor_name: str, action: str, entity: str = '', entity_id: str = '', meta: dict = None):
    await db.activity_logs.insert_one({
        '_id': new_id(),
        'actor_id': actor_id,
        'actor_name': actor_name,
        'action': action,
        'entity': entity,
        'entity_id': entity_id,
        'meta': meta or {},
        'timestamp': iso(now_utc()),
    })


# --------- Auth ---------
security = HTTPBearer(auto_error=False)


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    if not credentials or not credentials.credentials:
        raise HTTPException(status_code=401, detail='Not authenticated')
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail='Token expired')
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail='Invalid token')

    user = await db.users.find_one({'_id': payload['sub']})
    if not user:
        raise HTTPException(status_code=401, detail='User not found')
    if int(user.get('token_version', 0)) != int(payload.get('tv', 0)):
        raise HTTPException(status_code=401, detail='Sesi berakhir. Silakan login kembali.')
    user.pop('password_hash', None)
    user['id'] = user.pop('_id')
    roles = user_roles(user)
    tok_role = payload.get('role')
    user['roles'] = roles
    user['role'] = tok_role if tok_role in roles else (roles[0] if roles else user.get('role'))
    return user


def user_roles(u: dict) -> list:
    roles = u.get('roles')
    if roles:
        return [r for r in roles if r in ('admin', 'pengurus', 'peserta')]
    r = u.get('role')
    return [r] if r else []


def require_roles(*roles):
    async def _dep(user: dict = Depends(get_current_user)):
        if user.get('role') not in roles:
            raise HTTPException(status_code=403, detail='Forbidden')
        return user
    return _dep


# --------- Models (request/response) ---------
class LoginReq(BaseModel):
    identifier: str  # email OR phone OR username
    password: str
    role: Optional[Literal['admin', 'pengurus', 'peserta']] = None


class RegisterReq(BaseModel):
    name: str
    email: Optional[EmailStr] = None
    username: Optional[str] = None
    phone: Optional[str] = None
    password: str
    gender: Optional[Literal['L', 'P']] = 'L'
    birth_place: Optional[str] = ''
    birth_date: Optional[str] = ''
    education: Optional[str] = ''


class UserOut(BaseModel):
    model_config = ConfigDict(extra='ignore')
    id: str
    email: Optional[str] = None
    username: Optional[str] = None
    phone: Optional[str] = None
    name: str
    role: str
    participant_id: Optional[str] = None


class ChangePasswordReq(BaseModel):
    current_password: str
    new_password: str


class UserCreate(BaseModel):
    name: str
    email: Optional[EmailStr] = None
    username: Optional[str] = None
    phone: Optional[str] = None
    password: str
    role: Literal['admin', 'pengurus', 'peserta']
    participant_id: Optional[str] = None


class UserRoleUpdate(BaseModel):
    role: Optional[Literal['admin', 'pengurus', 'peserta']] = None
    roles: Optional[List[Literal['admin', 'pengurus', 'peserta']]] = None


class ParticipantCreate(BaseModel):
    name: str
    gender: Literal['L', 'P']
    birth_place: Optional[str] = ''
    birth_date: Optional[str] = ''  # ISO date
    phone: Optional[str] = ''
    email: Optional[str] = ''
    education: Optional[str] = ''
    status: Literal['aktif', 'non-aktif', 'arsip'] = 'aktif'
    is_secret_tag: bool = False
    duplicate_action: Optional[Literal['append', 'reject', 'allow']] = 'append'  # how to handle duplicate name


class ParticipantUpdate(BaseModel):
    name: Optional[str] = None
    gender: Optional[Literal['L', 'P']] = None
    birth_place: Optional[str] = None
    birth_date: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    education: Optional[str] = None
    status: Optional[Literal['aktif', 'non-aktif', 'arsip']] = None
    is_secret_tag: Optional[bool] = None


class ActivityCreate(BaseModel):
    name: str
    type: Literal['pengajian_rutin', 'pengajian_khusus', 'asad']
    date: str  # ISO date
    start_time: str  # HH:MM
    end_time: str
    location: Optional[str] = 'Kertalangu'
    gps_lat: Optional[float] = None
    gps_lng: Optional[float] = None
    radius_m: Optional[int] = 100
    is_outside: bool = False
    is_secret: bool = False
    pengajar: Optional[str] = ''
    materi_progress: Optional[str] = ''
    notes: Optional[str] = ''


class ActivityUpdate(BaseModel):
    name: Optional[str] = None
    type: Optional[Literal['pengajian_rutin', 'pengajian_khusus', 'asad']] = None
    date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    location: Optional[str] = None
    gps_lat: Optional[float] = None
    gps_lng: Optional[float] = None
    radius_m: Optional[int] = None
    is_outside: Optional[bool] = None
    is_secret: Optional[bool] = None
    pengajar: Optional[str] = None
    materi_progress: Optional[str] = None
    notes: Optional[str] = None
    manual_finished: Optional[bool] = None


class AttendanceScan(BaseModel):
    activity_id: str
    participant_qr: str  # QR payload from participant
    status: Optional[Literal['hadir', 'izin', 'alpha']] = 'hadir'


class AttendanceManual(BaseModel):
    activity_id: str
    participant_id: str
    status: Literal['hadir', 'izin', 'alpha']
    time_in: Optional[str] = None  # HH:MM, ignored for izin/alpha
    note: Optional[str] = ''


class MusyawarahCreate(BaseModel):
    kind: Literal['4S', 'TIM7']
    title: str
    content: str = ''
    date: Optional[str] = None


class MusyawarahUpdate(BaseModel):
    title: Optional[str] = None
    content: Optional[str] = None
    date: Optional[str] = None


class ReminderReq(BaseModel):
    activity_id: str
    participant_ids: List[str]


# --------- App ---------
app = FastAPI(title='E-Kertalangu API')
api = APIRouter(prefix='/api')


@api.get('/')
async def root():
    return {'app': 'E-Kertalangu', 'ok': True}


@api.get('/config/public')
async def public_config():
    return {
        'admin_wa': ADMIN_CONTACT_WA,
        'app_name': 'E-Kertalangu',
    }


# --------- Auth endpoints ---------
@api.post('/auth/login')
async def login(body: LoginReq, request: Request):
    ident = (body.identifier or '').strip()
    if not ident:
        raise HTTPException(status_code=400, detail='Identifier is required')
    email = ident.lower() if '@' in ident else None
    phone = normalize_phone(ident) if ident and ident[0].isdigit() or ident.startswith('+') else None
    username = ident.lower() if not email and not phone else None

    q = {'$or': []}
    if email:
        q['$or'].append({'email': email})
    if phone:
        q['$or'].append({'phone': phone})
    if username:
        q['$or'].append({'username': username})
    if not q['$or']:
        q = {'$or': [{'email': ident.lower()}, {'username': ident.lower()}, {'phone': normalize_phone(ident) or ident}]}

    user = await db.users.find_one(q)
    ip = request.client.host if request.client else 'unknown'
    key = ident.lower()

    # Pre-check lockout dalam jendela waktu (SEC-005)
    att = await db.login_attempts.find_one({'_id': key})
    if att and att.get('count', 0) >= LOGIN_MAX_ATTEMPTS:
        try:
            first = datetime.fromisoformat(att.get('first_at')) if att.get('first_at') else now_utc()
        except Exception:
            first = now_utc()
        elapsed = (now_utc() - first).total_seconds()
        if elapsed < LOGIN_LOCK_WINDOW_SEC:
            mins = max(1, int((LOGIN_LOCK_WINDOW_SEC - elapsed) // 60) + 1)
            raise HTTPException(status_code=429, detail=f'Terlalu banyak percobaan. Coba lagi dalam {mins} menit.')
        # jendela lewat -> reset
        await db.login_attempts.delete_one({'_id': key})
        att = None

    if not user or not verify_pw(body.password, user.get('password_hash', '')):
        now_iso = iso(now_utc())
        if att:
            await db.login_attempts.update_one({'_id': key}, {'$inc': {'count': 1}, '$set': {'ts': now_iso}})
        else:
            await db.login_attempts.update_one({'_id': key}, {'$set': {'count': 1, 'first_at': now_iso, 'ts': now_iso}}, upsert=True)
        raise HTTPException(status_code=401, detail='Kredensial salah')

    if user.get('active') is False:
        if user.get('pending_approval'):
            raise HTTPException(status_code=403, detail='Akun Anda menunggu persetujuan admin. Silakan hubungi admin.')
        raise HTTPException(status_code=403, detail='Akun dinonaktifkan')

    # Blokir login jika peserta terkait diarsipkan (Batch 4)
    if user.get('participant_id'):
        par = await db.participants.find_one({'_id': user['participant_id']}, {'status': 1})
        if par and par.get('status') == 'arsip':
            raise HTTPException(status_code=403, detail='Akun Anda telah diarsipkan. Silakan hubungi admin.')

    await db.login_attempts.delete_one({'_id': key})

    roles = user_roles(user) or ['peserta']
    if body.role:
        if body.role not in roles:
            raise HTTPException(status_code=403, detail='Role tersebut tidak tersedia untuk akun ini')
        chosen = body.role
    elif len(roles) > 1:
        # Multiple roles: client must pick one (poin 12)
        return {'needs_role': True, 'roles': roles, 'name': user.get('name', '')}
    else:
        chosen = roles[0]

    token = make_token(user['_id'], chosen, int(user.get('token_version', 0)))

    await log_activity(user['_id'], user.get('name', ''), 'login', 'user', user['_id'], {'ip': ip, 'role': chosen})

    user_out = {k: v for k, v in user.items() if k != 'password_hash'}
    user_out['id'] = user_out.pop('_id')
    user_out['role'] = chosen
    user_out['roles'] = roles
    return {'token': token, 'user': user_out}


@api.post('/auth/register')
async def register(body: RegisterReq, request: Request):
    name = (body.name or '').strip()
    if not name:
        raise HTTPException(status_code=400, detail='Nama wajib diisi')
    if len(body.password) < 4:
        raise HTTPException(status_code=400, detail='Password minimal 4 karakter')
    # Warning flag (soft) — password lemah tapi tetap diterima
    password_warning = None
    if len(body.password) < 6:
        password_warning = 'Password Anda pendek. Untuk keamanan lebih baik, gunakan minimal 6 karakter.'
    email = (body.email or '').lower() or None
    username = (body.username or '').lower() or None
    phone = normalize_phone(body.phone) or None
    if not (email or username or phone):
        raise HTTPException(status_code=400, detail='Isi salah satu: email, username, atau nomor HP')
    or_q = []
    if email: or_q.append({'email': email})
    if username: or_q.append({'username': username})
    if phone: or_q.append({'phone': phone})
    existing = await db.users.find_one({'$or': or_q}) if or_q else None
    if existing:
        raise HTTPException(status_code=400, detail='Akun dengan email/username/HP tersebut sudah ada')

    # Auto-create linked participant record
    code = await _next_participant_code()
    qr_payload = f'EKTL:P:{code}'
    part_id = new_id()
    await db.participants.insert_one({
        '_id': part_id,
        'code': code,
        'name': name,
        'name_key': name.lower(),
        'gender': body.gender or 'L',
        'birth_place': body.birth_place or '',
        'birth_date': body.birth_date or '',
        'phone': phone or '',
        'email': email or '',
        'education': body.education or '',
        'status': 'aktif',
        'is_secret_tag': False,
        'qr_payload': qr_payload,
        'created_at': iso(now_utc()),
    })

    user_id = new_id()
    await db.users.insert_one({
        '_id': user_id,
        'email': email,
        'username': username,
        'phone': phone,
        'name': name,
        'password_hash': hash_pw(body.password),
        'role': 'peserta',
        'roles': ['peserta'],
        'participant_id': part_id,
        'active': True,  # tanpa persetujuan admin — langsung aktif
        'pending_approval': False,
        'created_at': iso(now_utc()),
    })

    await log_activity(user_id, name, 'register', 'user', user_id, {'via': 'self-register', 'pending': False})
    token = make_token(user_id, 'peserta')
    user_out = {
        'id': user_id, 'email': email, 'username': username, 'phone': phone, 'name': name,
        'role': 'peserta', 'participant_id': part_id, 'active': True, 'pending_approval': False,
    }
    return {
        'pending': False,
        'token': token,
        'message': 'Pendaftaran berhasil! Anda langsung bisa masuk.',
        'password_warning': password_warning,
        'user': user_out,
    }


@api.get('/auth/me')
async def me(user: dict = Depends(get_current_user)):
    return {'user': user}


# --------- Activation (peserta hasil import melengkapi akun) ---------
class ActivationActivate(BaseModel):
    participant_id: str
    phone: str
    password: str


@api.get('/activation/search')
async def activation_search(q: str):
    q = (q or '').strip()
    if len(q) < 3:
        return {'items': []}
    rx = {'$regex': re.escape(q), '$options': 'i'}
    linked_pids = set()
    async for u in db.users.find({'participant_id': {'$ne': None}}, {'participant_id': 1}):
        if u.get('participant_id'):
            linked_pids.add(u['participant_id'])
    items = []
    cur = db.participants.find({'$or': [{'name': rx}, {'code': rx}]}).sort('name', 1).limit(20)
    async for p in cur:
        if p['_id'] in linked_pids:
            continue  # sudah punya akun
        if p.get('is_secret_tag'):
            continue
        ph = p.get('phone') or ''
        masked = (ph[:4] + '****' + ph[-2:]) if len(ph) >= 6 else ('' if not ph else ph)
        items.append({'id': p['_id'], 'code': p['code'], 'name': p['name'], 'phone_masked': masked, 'has_phone': bool(ph)})
    return {'items': items}


@api.post('/activation/activate')
async def activation_activate(body: ActivationActivate):
    p = await db.participants.find_one({'_id': body.participant_id})
    if not p:
        raise HTTPException(status_code=404, detail='Data peserta tidak ditemukan')
    # Cegah duplikasi: satu peserta tidak boleh terhubung ke dua akun
    existing_link = await db.users.find_one({'participant_id': body.participant_id})
    if existing_link:
        raise HTTPException(status_code=409, detail='Peserta ini sudah memiliki akun. Silakan login.')
    phone = normalize_phone(body.phone)
    if not phone:
        raise HTTPException(status_code=400, detail='Nomor HP wajib diisi')
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail='Password minimal 6 karakter')
    # Verifikasi identitas: jika peserta SUDAH punya No. HP terdaftar, nomor harus cocok (SEC-002).
    # Jika BELUM ada No. HP terdaftar, izinkan aktivasi mandiri dengan nomor yang dimasukkan (disimpan sbg No. HP peserta).
    registered_phone = normalize_phone(p.get('phone') or '')
    if registered_phone and phone != registered_phone:
        raise HTTPException(status_code=400, detail='Nomor HP tidak cocok dengan data yang terdaftar. Hubungi pengurus jika ada perubahan nomor.')
    if await db.users.find_one({'phone': phone}):
        raise HTTPException(status_code=400, detail='Nomor HP sudah digunakan akun lain')
    user_id = new_id()
    await db.users.insert_one({
        '_id': user_id,
        'email': (p.get('email') or None),
        'username': None,
        'phone': phone,
        'name': p['name'],
        'password_hash': hash_pw(body.password),
        'role': 'peserta',
        'roles': ['peserta'],
        'participant_id': body.participant_id,  # TAUTKAN ke participant yang sudah ada (tanpa duplikat)
        'active': True,
        'pending_approval': False,
        'created_at': iso(now_utc()),
    })
    upd = {'status': 'aktif'}
    if not p.get('phone'):
        upd['phone'] = phone
    await db.participants.update_one({'_id': body.participant_id}, {'$set': upd})
    await log_activity(user_id, p['name'], 'activate', 'user', user_id, {'participant_id': body.participant_id})
    token = make_token(user_id, 'peserta')
    return {
        'token': token,
        'user': {
            'id': user_id, 'name': p['name'], 'phone': phone, 'role': 'peserta',
            'roles': ['peserta'], 'participant_id': body.participant_id, 'active': True,
        },
    }



@api.post('/auth/logout')
async def logout(user: dict = Depends(get_current_user)):
    # Cabut semua token lama (SEC-004)
    await db.users.update_one({'_id': user['id']}, {'$inc': {'token_version': 1}})
    await log_activity(user['id'], user.get('name', ''), 'logout', 'user', user['id'])
    return {'ok': True}


@api.post('/auth/change-password')
async def change_password(body: ChangePasswordReq, user: dict = Depends(get_current_user)):
    dbu = await db.users.find_one({'_id': user['id']})
    if not dbu or not verify_pw(body.current_password, dbu.get('password_hash', '')):
        raise HTTPException(status_code=400, detail='Password saat ini salah')
    if len(body.new_password) < 6:
        raise HTTPException(status_code=400, detail='Password baru minimal 6 karakter')
    new_tv = int(dbu.get('token_version', 0)) + 1
    await db.users.update_one({'_id': user['id']}, {'$set': {'password_hash': hash_pw(body.new_password), 'token_version': new_tv}})
    await log_activity(user['id'], user.get('name', ''), 'change_password')
    # Re-issue token utk sesi saat ini agar tetap valid; sesi lain tercabut
    token = make_token(user['id'], user.get('role'), new_tv)
    return {'ok': True, 'token': token}


# --------- User management (admin) ---------
@api.get('/auth/pending-count')
async def pending_count(user: dict = Depends(require_roles('admin'))):
    n = await db.users.count_documents({'pending_approval': True, 'active': False})
    return {'count': n}


@api.post('/users/{uid}/approve')
async def approve_user(uid: str, user: dict = Depends(require_roles('admin'))):
    r = await db.users.update_one({'_id': uid}, {'$set': {'active': True, 'pending_approval': False}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail='Pengguna tidak ditemukan')
    await log_activity(user['id'], user.get('name', ''), 'approve_user', 'user', uid)
    return {'ok': True}


@api.post('/users/{uid}/reject')
async def reject_user(uid: str, user: dict = Depends(require_roles('admin'))):
    u = await db.users.find_one({'_id': uid})
    if not u:
        raise HTTPException(status_code=404, detail='Pengguna tidak ditemukan')
    await db.users.delete_one({'_id': uid})
    if u.get('participant_id'):
        await db.participants.delete_one({'_id': u['participant_id']})
    await log_activity(user['id'], user.get('name', ''), 'reject_user', 'user', uid)
    return {'ok': True}


@api.get('/users')
async def list_users(user: dict = Depends(require_roles('admin'))):
    cur = db.users.find({}, {'password_hash': 0}).sort('name', 1)
    out = []
    async for u in cur:
        u['id'] = u.pop('_id')
        u['roles'] = user_roles(u)
        out.append(u)
    return {'users': out}


@api.post('/users')
async def create_user(body: UserCreate, user: dict = Depends(require_roles('admin'))):
    email = (body.email or '').lower() or None
    username = (body.username or '').lower() or None
    phone = normalize_phone(body.phone) or None
    if not (email or username or phone):
        raise HTTPException(status_code=400, detail='Isi salah satu: email, username, atau nomor HP')
    existing = await db.users.find_one({'$or': [
        {'email': email} if email else {'_never': True},
        {'username': username} if username else {'_never': True},
        {'phone': phone} if phone else {'_never': True},
    ]})
    if existing:
        raise HTTPException(status_code=400, detail='Pengguna sudah ada (email/username/phone duplikat)')
    doc = {
        '_id': new_id(),
        'email': email,
        'username': username,
        'phone': phone,
        'name': body.name,
        'password_hash': hash_pw(body.password),
        'role': body.role,
        'roles': [body.role],
        'participant_id': body.participant_id,
        'active': True,
        'created_at': iso(now_utc()),
    }
    await db.users.insert_one(doc)
    await log_activity(user['id'], user.get('name', ''), 'create_user', 'user', doc['_id'], {'role': body.role})
    return {'ok': True, 'id': doc['_id']}


@api.patch('/users/{uid}/role')
async def update_role(uid: str, body: UserRoleUpdate, user: dict = Depends(require_roles('admin'))):
    incoming = body.roles if body.roles is not None else ([body.role] if body.role else None)
    if not incoming:
        raise HTTPException(status_code=400, detail='Pilih minimal satu role')
    roles = []
    for r in incoming:
        if r not in roles:
            roles.append(r)
    r = await db.users.update_one({'_id': uid}, {'$set': {'roles': roles, 'role': roles[0]}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail='Pengguna tidak ditemukan')
    target = await db.users.find_one({'_id': uid}, {'name': 1})
    await log_activity(user['id'], user.get('name', ''), 'update_role', 'user', uid, {'target': (target or {}).get('name', ''), 'roles': roles})
    return {'ok': True, 'roles': roles}


@api.patch('/users/{uid}/toggle-active')
async def toggle_active(uid: str, user: dict = Depends(require_roles('admin'))):
    u = await db.users.find_one({'_id': uid})
    if not u:
        raise HTTPException(status_code=404, detail='Pengguna tidak ditemukan')
    new_active = not bool(u.get('active', True))
    await db.users.update_one({'_id': uid}, {'$set': {'active': new_active}})
    await log_activity(user['id'], user.get('name', ''), 'toggle_active', 'user', uid, {'target': u.get('name', ''), 'active': new_active})
    return {'ok': True, 'active': new_active}


# --------- Participants ---------
async def _next_participant_code() -> str:
    counter = await db.counters.find_one_and_update(
        {'_id': 'participant'},
        {'$inc': {'seq': 1}},
        upsert=True,
        return_document=True,
    )
    seq = counter.get('seq', 1) if counter else 1
    return f'KTL-{seq:04d}'


@api.get('/participants')
async def list_participants(
    q: Optional[str] = None,
    status_f: Optional[str] = Query(None, alias='status'),
    limit: int = 500,
    user: dict = Depends(require_roles('admin', 'pengurus')),
):
    filt: dict = {}
    if q and len(q) >= 2:
        rx = {'$regex': re.escape(q), '$options': 'i'}
        filt['$or'] = [
            {'name': rx},
            {'code': rx},
            {'phone': rx},
            {'email': rx},
        ]
    if status_f in ('aktif', 'non-aktif', 'arsip'):
        if status_f == 'arsip' and user.get('role') != 'admin':
            raise HTTPException(status_code=403, detail='Hanya admin yang dapat melihat data arsip')
        filt['status'] = status_f
    else:
        # Default: sembunyikan peserta arsip dari daftar aktif (Batch 4)
        filt['status'] = {'$ne': 'arsip'}
    cur = db.participants.find(filt).sort('name', 1).limit(limit)
    items = []
    async for p in cur:
        p['id'] = p.pop('_id')
        # hide secret tagged from peserta/pengurus? Only admin can see secret detail
        if p.get('is_secret_tag') and user.get('role') == 'peserta':
            continue
        items.append(p)
    # Derive account status (single source of truth: user.participant_id link)
    pids = [p['id'] for p in items]
    linked = set()
    if pids:
        async for u in db.users.find({'participant_id': {'$in': pids}}, {'participant_id': 1}):
            if u.get('participant_id'):
                linked.add(u['participant_id'])
    for p in items:
        p['account_status'] = 'aktif' if p['id'] in linked else 'belum_aktivasi'
    return {'items': items}


@api.post('/participants')
async def create_participant(body: ParticipantCreate, user: dict = Depends(require_roles('admin', 'pengurus'))):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail='Nama wajib diisi')
    dup = await db.participants.count_documents({'name_key': name.lower()})
    action = body.duplicate_action or 'append'
    if dup and action == 'reject':
        raise HTTPException(status_code=409, detail='Nama sudah ada. Gunakan opsi lain.')

    final_name = name
    if dup and action == 'append':
        final_name = f'{name} ({dup + 1})'

    code = await _next_participant_code()
    qr_payload = f'EKTL:P:{code}'
    doc = {
        '_id': new_id(),
        'code': code,
        'name': final_name,
        'name_key': final_name.lower(),
        'gender': body.gender,
        'birth_place': body.birth_place or '',
        'birth_date': body.birth_date or '',
        'phone': normalize_phone(body.phone) or '',
        'email': (body.email or '').lower(),
        'education': body.education or '',
        'status': body.status,
        'is_secret_tag': bool(body.is_secret_tag),
        'qr_payload': qr_payload,
        'created_at': iso(now_utc()),
    }
    await db.participants.insert_one(doc)
    await log_activity(user['id'], user.get('name', ''), 'create_participant', 'participant', doc['_id'], {'name': final_name})
    doc['id'] = doc.pop('_id')
    return doc


@api.post('/participants/bulk')
async def bulk_create(items: List[ParticipantCreate], user: dict = Depends(require_roles('admin', 'pengurus'))):
    results = []
    for it in items:
        try:
            r = await create_participant(it, user)  # type: ignore
            results.append({'ok': True, 'id': r['id'], 'name': r['name']})
        except HTTPException as e:
            results.append({'ok': False, 'error': e.detail, 'name': it.name})
    return {'results': results}


@api.get('/participants/{pid}')
async def get_participant(pid: str, user: dict = Depends(get_current_user)):
    if user.get('role') not in ('admin', 'pengurus') and user.get('participant_id') != pid:
        raise HTTPException(status_code=403, detail='Anda tidak memiliki akses ke data peserta ini')
    p = await db.participants.find_one({'_id': pid})
    if not p:
        raise HTTPException(status_code=404, detail='Peserta tidak ditemukan')
    p['id'] = p.pop('_id')
    p['qr_datauri'] = qr_png_datauri(p['qr_payload'])
    return p


@api.get('/participants/{pid}/account')
async def get_participant_account(pid: str, user: dict = Depends(require_roles('admin'))):
    p = await db.participants.find_one({'_id': pid}, {'_id': 1})
    if not p:
        raise HTTPException(status_code=404, detail='Peserta tidak ditemukan')
    u = await db.users.find_one({'participant_id': pid}, {'password_hash': 0})
    if not u:
        return {'has_account': False, 'user_id': None, 'roles': []}
    return {'has_account': True, 'user_id': u['_id'], 'roles': user_roles(u), 'active': bool(u.get('active', True))}


@api.patch('/participants/{pid}')
async def update_participant(pid: str, body: ParticipantUpdate, user: dict = Depends(require_roles('admin', 'pengurus'))):
    upd = {k: v for k, v in body.model_dump().items() if v is not None}
    if 'name' in upd:
        upd['name_key'] = upd['name'].lower()
    if 'phone' in upd:
        upd['phone'] = normalize_phone(upd['phone']) or ''
    if 'email' in upd:
        upd['email'] = (upd['email'] or '').lower()
    r = await db.participants.update_one({'_id': pid}, {'$set': upd})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail='Peserta tidak ditemukan')
    await log_activity(user['id'], user.get('name', ''), 'update_participant', 'participant', pid, upd)
    return {'ok': True}


@api.delete('/participants/{pid}')
async def delete_participant(pid: str, user: dict = Depends(require_roles('admin'))):
    p = await db.participants.find_one({'_id': pid}, {'name': 1})
    r = await db.participants.delete_one({'_id': pid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Peserta tidak ditemukan')
    # Hapus akun login tertaut (data historis absensi tetap dipertahankan)
    await db.users.delete_many({'participant_id': pid})
    await log_activity(user['id'], user.get('name', ''), 'delete_participant', 'participant', pid, {'name': (p or {}).get('name', '')})
    return {'ok': True}


@api.patch('/participants/{pid}/archive')
async def archive_participant(pid: str, user: dict = Depends(require_roles('admin'))):
    p = await db.participants.find_one({'_id': pid}, {'name': 1})
    if not p:
        raise HTTPException(status_code=404, detail='Peserta tidak ditemukan')
    await db.participants.update_one({'_id': pid}, {'$set': {'status': 'arsip'}})
    # Cabut sesi akun tertaut agar tidak bisa dipakai lagi
    await db.users.update_many({'participant_id': pid}, {'$inc': {'token_version': 1}})
    await log_activity(user['id'], user.get('name', ''), 'archive_participant', 'participant', pid, {'name': p.get('name', '')})
    return {'ok': True}


@api.patch('/participants/{pid}/restore')
async def restore_participant(pid: str, user: dict = Depends(require_roles('admin'))):
    p = await db.participants.find_one({'_id': pid}, {'name': 1})
    if not p:
        raise HTTPException(status_code=404, detail='Peserta tidak ditemukan')
    await db.participants.update_one({'_id': pid}, {'$set': {'status': 'aktif'}})
    await log_activity(user['id'], user.get('name', ''), 'restore_participant', 'participant', pid, {'name': p.get('name', '')})
    return {'ok': True}


class BulkIdsReq(BaseModel):
    ids: List[str]


@api.post('/participants/bulk-delete')
async def bulk_delete_participants(body: BulkIdsReq, user: dict = Depends(require_roles('admin'))):
    ids = [i for i in (body.ids or []) if i]
    if not ids:
        raise HTTPException(status_code=400, detail='Tidak ada peserta yang dipilih')
    names = []
    async for p in db.participants.find({'_id': {'$in': ids}}, {'name': 1}):
        names.append(p.get('name', ''))
    r = await db.participants.delete_many({'_id': {'$in': ids}})
    await db.users.delete_many({'participant_id': {'$in': ids}})
    await log_activity(user['id'], user.get('name', ''), 'delete_participant', 'participant', '', {'bulk': True, 'count': r.deleted_count, 'names': names[:50]})
    return {'ok': True, 'deleted': r.deleted_count}


@api.post('/participants/{pid}/reset-password')
async def reset_participant_password(pid: str, user: dict = Depends(require_roles('admin'))):
    p = await db.participants.find_one({'_id': pid}, {'name': 1})
    if not p:
        raise HTTPException(status_code=404, detail='Peserta tidak ditemukan')
    linked = await db.users.find_one({'participant_id': pid})
    if not linked:
        raise HTTPException(status_code=400, detail='Peserta belum memiliki akun (belum aktivasi)')
    if 'admin' in user_roles(linked):
        raise HTTPException(status_code=400, detail='Tidak dapat mereset password akun admin dari sini')
    temp_pw = _gen_temp_password()
    new_tv = int(linked.get('token_version', 0)) + 1
    await db.users.update_one({'_id': linked['_id']}, {'$set': {'password_hash': hash_pw(temp_pw), 'token_version': new_tv}})
    await log_activity(user['id'], user.get('name', ''), 'reset_participant_password', 'participant', pid, {'name': p.get('name', '')})
    return {'ok': True, 'temp_password': temp_pw, 'participant_name': p.get('name', '')}


@api.get('/participants/{pid}/attendance')
async def participant_attendance(pid: str, user: dict = Depends(get_current_user)):
    if user.get('role') not in ('admin', 'pengurus') and user.get('participant_id') != pid:
        raise HTTPException(status_code=403, detail='Anda tidak memiliki akses ke data peserta ini')
    rows = await db.attendance.find({'participant_id': pid}).sort('recorded_at', -1).limit(500).to_list(500)
    act_ids = list({r['activity_id'] for r in rows if r.get('activity_id')})
    acts = {}
    if act_ids:
        async for act in db.activities.find({'_id': {'$in': act_ids}}):
            acts[act['_id']] = act
    items = []
    for a in rows:
        act = acts.get(a.get('activity_id'))
        a['id'] = a.pop('_id')
        a['activity'] = {'id': act['_id'], 'name': act['name'], 'date': act.get('date', ''), 'type': act.get('type', '')} if act else None
        items.append(a)
    return {'items': items}


@api.get('/participants/{pid}/stats')
async def participant_stats(pid: str, user: dict = Depends(get_current_user)):
    if user.get('role') not in ('admin', 'pengurus') and user.get('participant_id') != pid:
        raise HTTPException(status_code=403, detail='Anda tidak memiliki akses ke data peserta ini')
    pipe = [
        {'$match': {'participant_id': pid}},
        {'$group': {'_id': '$status', 'n': {'$sum': 1}}},
    ]
    agg = {r['_id']: r['n'] async for r in db.attendance.aggregate(pipe)}
    total = sum(agg.values()) or 1
    return {
        'counts': agg,
        'rate_hadir': round((agg.get('hadir', 0) / total) * 100, 1),
        'total': sum(agg.values()),
    }


# --------- Activities ---------
async def _next_activity_code() -> str:
    counter = await db.counters.find_one_and_update(
        {'_id': 'activity'},
        {'$inc': {'seq': 1}},
        upsert=True,
        return_document=True,
    )
    seq = counter.get('seq', 1) if counter else 1
    return f'ACT-{seq:04d}'


@api.get('/activities')
async def list_activities(
    q: Optional[str] = None,
    type_f: Optional[str] = Query(None, alias='type'),
    upcoming: bool = False,
    user: dict = Depends(get_current_user),
):
    filt: dict = {}
    if q:
        filt['name'] = {'$regex': re.escape(q), '$options': 'i'}
    if type_f:
        filt['type'] = type_f
    if upcoming:
        filt['date'] = {'$gte': datetime.now().strftime('%Y-%m-%d')}
    # Peserta cannot see secret activities unless tagged
    cur = db.activities.find(filt).sort([('date', -1), ('start_time', -1)]).limit(500)
    items = []
    async for a in cur:
        if a.get('is_secret') and user.get('role') == 'peserta':
            # Only show if user's participant is in secret allowlist
            allowed = a.get('secret_allow', [])
            if user.get('participant_id') not in allowed:
                continue
        a['id'] = a.pop('_id')
        items.append(a)
    return {'items': items}


@api.post('/activities')
async def create_activity(body: ActivityCreate, user: dict = Depends(require_roles('admin', 'pengurus'))):
    code = await _next_activity_code()
    qr_payload = f'EKTL:A:{code}'
    doc = {
        '_id': new_id(),
        'code': code,
        'name': body.name,
        'type': body.type,
        'date': body.date,
        'start_time': body.start_time,
        'end_time': body.end_time,
        'location': body.location or 'Kertalangu',
        'gps_lat': body.gps_lat,
        'gps_lng': body.gps_lng,
        'radius_m': int(body.radius_m or 100),
        'is_outside': bool(body.is_outside),
        'is_secret': bool(body.is_secret),
        'secret_allow': [],
        'pengajar': body.pengajar or '',
        'materi_progress': body.materi_progress or '',
        'notes': body.notes or '',
        'qr_payload': qr_payload,
        'created_at': iso(now_utc()),
        'created_by': user['id'],
    }
    await db.activities.insert_one(doc)
    await log_activity(user['id'], user.get('name', ''), 'create_activity', 'activity', doc['_id'], {'name': body.name})
    await _notify_activity(doc)
    doc['id'] = doc.pop('_id')
    return doc


@api.get('/activities/{aid}')
async def get_activity(aid: str, user: dict = Depends(get_current_user)):
    a = await db.activities.find_one({'_id': aid})
    if not a:
        raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan')
    # Kegiatan rahasia: peserta hanya boleh membuka bila ada di allowlist (SEC-001)
    if a.get('is_secret') and user.get('role') == 'peserta':
        if user.get('participant_id') not in (a.get('secret_allow') or []):
            raise HTTPException(status_code=403, detail='Tidak diizinkan mengakses kegiatan ini')
    a['id'] = a.pop('_id')
    a['qr_datauri'] = qr_png_datauri(a['qr_payload'])
    return a


@api.patch('/activities/{aid}')
async def update_activity(aid: str, body: ActivityUpdate, user: dict = Depends(require_roles('admin', 'pengurus'))):
    upd = {k: v for k, v in body.model_dump().items() if v is not None}
    r = await db.activities.update_one({'_id': aid}, {'$set': upd})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan')
    await log_activity(user['id'], user.get('name', ''), 'update_activity', 'activity', aid, upd)
    return {'ok': True}


@api.delete('/activities/{aid}')
async def delete_activity(aid: str, user: dict = Depends(require_roles('admin'))):
    r = await db.activities.delete_one({'_id': aid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan')
    await db.attendance.delete_many({'activity_id': aid})
    await log_activity(user['id'], user.get('name', ''), 'delete_activity', 'activity', aid)
    return {'ok': True}


class SecretAllow(BaseModel):
    participant_ids: List[str]


@api.post('/activities/{aid}/secret-allow')
async def set_secret_allow(aid: str, body: SecretAllow, user: dict = Depends(require_roles('admin', 'pengurus'))):
    r = await db.activities.update_one({'_id': aid}, {'$set': {'secret_allow': body.participant_ids, 'is_secret': True}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan')
    return {'ok': True}


# --------- Notifications (in-app) ---------
async def _notify_activity(activity: dict):
    try:
        title = 'Kegiatan Baru'
        body = f"{activity.get('name', '')} · {activity.get('date', '')} {activity.get('start_time', '')} WITA @ {activity.get('location', '')}"
        now = iso(now_utc())
        docs = []
        cur = db.users.find({'active': {'$ne': False}}, {'_id': 1, 'roles': 1, 'role': 1})
        async for u in cur:
            roles = u.get('roles') or ([u.get('role')] if u.get('role') else [])
            if 'peserta' in roles or 'pengurus' in roles:
                docs.append({
                    '_id': new_id(),
                    'user_id': u['_id'],
                    'title': title,
                    'body': body,
                    'type': 'activity',
                    'activity_id': activity.get('_id'),
                    'read': False,
                    'created_at': now,
                })
        if docs:
            await db.notifications.insert_many(docs)
    except Exception as e:
        logger.warning(f'notify failed: {e}')


@api.get('/notifications')
async def list_notifications(user: dict = Depends(get_current_user)):
    cur = db.notifications.find({'user_id': user['id']}).sort('created_at', -1).limit(50)
    items = []
    unread = 0
    async for n in cur:
        n['id'] = n.pop('_id')
        if not n.get('read'):
            unread += 1
        items.append(n)
    return {'items': items, 'unread': unread}


@api.post('/notifications/{nid}/read')
async def read_notification(nid: str, user: dict = Depends(get_current_user)):
    await db.notifications.update_one({'_id': nid, 'user_id': user['id']}, {'$set': {'read': True}})
    return {'ok': True}


@api.post('/notifications/read-all')
async def read_all_notifications(user: dict = Depends(get_current_user)):
    await db.notifications.update_many({'user_id': user['id'], 'read': {'$ne': True}}, {'$set': {'read': True}})
    return {'ok': True}


# --------- Attendance ---------
def _activity_finished(act: dict) -> bool:
    """Kegiatan dianggap selesai jika ditutup manual (manual_finished True),
    atau (jika belum diatur manual) waktu server sudah melewati jam selesai.
    manual_finished False = dibuka kembali (override, tidak terkunci)."""
    mf = act.get('manual_finished')
    if mf is True:
        return True
    if mf is False:
        return False
    try:
        end = datetime.strptime(f"{act.get('date')} {act.get('end_time')}", '%Y-%m-%d %H:%M')
    except Exception:
        return False
    return now_wita() > end


async def _resolve_self_participant(user: dict) -> Optional[dict]:
    """Resolusi data peserta untuk absen mandiri.
    Prioritas identifier existing (ID-based), TANPA membuat peserta baru dan TANPA
    mengandalkan nama sebagai identitas utama. Jika tautan akun putus (mis. peserta
    di-import ulang sehingga _id berubah), cari ulang via identifier unik yang sudah ada
    lalu perbaiki tautan akun ini saja (bukan migrasi massal)."""
    pid = user.get('participant_id')
    if pid:
        par = await db.participants.find_one({'_id': pid})
        if par:
            return par
    # Fallback: identifier unik existing (No. HP, lalu email)
    queries = []
    uphone = normalize_phone(user.get('phone'))
    if uphone:
        queries.append(('phone', {'phone': uphone}))
    uemail = (user.get('email') or '').strip().lower()
    if uemail:
        queries.append(('email', {'email': uemail}))
    for via, q in queries:
        par = await db.participants.find_one(q)
        if par:
            if par['_id'] != user.get('participant_id'):
                await db.users.update_one({'_id': user['id']}, {'$set': {'participant_id': par['_id']}})
                await log_activity(user['id'], user.get('name', ''), 'relink_participant', 'user', user['id'], {'participant_id': par['_id'], 'via': via})
            return par
    # Fallback terakhir: nama persis & UNIK (hindari salah cocok jika ada nama sama)
    uname = (user.get('name') or '').strip().lower()
    if uname:
        matches = await db.participants.find({'name_key': uname}).to_list(3)
        if len(matches) == 1:
            par = matches[0]
            if par['_id'] != user.get('participant_id'):
                await db.users.update_one({'_id': user['id']}, {'$set': {'participant_id': par['_id']}})
                await log_activity(user['id'], user.get('name', ''), 'relink_participant', 'user', user['id'], {'participant_id': par['_id'], 'via': 'name'})
            return par
    return None


async def _record_attendance(activity_id: str, participant_id: str, status_: str, time_in: Optional[str], method: str, note: str, recorder: dict):
    act = await db.activities.find_one({'_id': activity_id})
    if not act:
        raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan')
    if _activity_finished(act):
        raise HTTPException(status_code=400, detail='Kegiatan sudah selesai. Absensi ditutup.')
    par = await db.participants.find_one({'_id': participant_id})
    if not par:
        raise HTTPException(status_code=404, detail='Peserta tidak ditemukan')
    if par.get('status') == 'arsip':
        raise HTTPException(status_code=400, detail='Peserta telah diarsipkan, tidak dapat melakukan absensi.')
    tin = time_in
    if status_ == 'hadir' and not tin:
        tin = now_wita().strftime('%H:%M')

    existing = await db.attendance.find_one({'activity_id': activity_id, 'participant_id': participant_id})
    doc = {
        'activity_id': activity_id,
        'participant_id': participant_id,
        'participant_name': par['name'],
        'participant_code': par.get('code', ''),
        'activity_name': act['name'],
        'activity_date': act.get('date', ''),
        'status': status_,
        'time_in': tin if status_ == 'hadir' else '',
        'method': method,
        'note': note or '',
        'recorded_by': recorder['id'],
        'recorded_by_name': recorder.get('name', ''),
        'recorded_at': iso(now_utc()),
    }
    if existing:
        await db.attendance.update_one({'_id': existing['_id']}, {'$set': doc})
        aid = existing['_id']
    else:
        doc['_id'] = new_id()
        await db.attendance.insert_one(doc)
        aid = doc['_id']
    await log_activity(recorder['id'], recorder.get('name', ''), 'attendance', 'attendance', aid, {'status': status_, 'participant': par['name'], 'activity': act['name']})
    doc['id'] = aid
    doc.pop('_id', None)
    return doc


@api.post('/attendance/scan')
async def attendance_scan(body: AttendanceScan, user: dict = Depends(require_roles('admin', 'pengurus'))):
    payload = (body.participant_qr or '').strip()
    if not payload.startswith('EKTL:P:'):
        raise HTTPException(status_code=400, detail='QR tidak valid')
    code = payload.split('EKTL:P:', 1)[1]
    par = await db.participants.find_one({'code': code})
    if not par:
        raise HTTPException(status_code=404, detail='Peserta dari QR tidak ditemukan')
    return await _record_attendance(body.activity_id, par['_id'], body.status or 'hadir', None, 'qr', '', user)


@api.post('/attendance/manual')
async def attendance_manual(body: AttendanceManual, user: dict = Depends(require_roles('admin', 'pengurus'))):
    return await _record_attendance(body.activity_id, body.participant_id, body.status, body.time_in, 'manual', body.note or '', user)


@api.post('/activities/{aid}/mark-remaining-alpha')
async def mark_remaining_alpha(aid: str, user: dict = Depends(require_roles('admin', 'pengurus'))):
    """Tandai semua peserta aktif yang BELUM absen sebagai Alpha (finalisasi kegiatan).
    Tidak menimpa absensi yang sudah ada; bisa dipakai walau kegiatan sudah ditutup."""
    act = await db.activities.find_one({'_id': aid})
    if not act:
        raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan')
    existing = set(await db.attendance.distinct('participant_id', {'activity_id': aid}))
    count = 0
    cur = db.participants.find({'status': 'aktif'})
    async for p in cur:
        if p['_id'] in existing:
            continue
        await db.attendance.insert_one({
            '_id': new_id(),
            'activity_id': aid,
            'participant_id': p['_id'],
            'participant_name': p['name'],
            'participant_code': p.get('code', ''),
            'activity_name': act['name'],
            'activity_date': act.get('date', ''),
            'status': 'alpha',
            'time_in': '',
            'method': 'manual',
            'note': 'Tandai semua alpha',
            'recorded_by': user['id'],
            'recorded_by_name': user.get('name', ''),
            'recorded_at': iso(now_utc()),
        })
        count += 1
    await log_activity(user['id'], user.get('name', ''), 'mark_remaining_alpha', 'activity', aid, {'count': count})
    return {'ok': True, 'count': count}


@api.get('/attendance/by-activity/{aid}')
async def by_activity(aid: str, user: dict = Depends(require_roles('admin', 'pengurus'))):
    cur = db.attendance.find({'activity_id': aid}).sort('recorded_at', -1)
    items = []
    async for a in cur:
        a['id'] = a.pop('_id')
        items.append(a)
    # Daftar peserta (termasuk No. HP untuk pengingat) hanya untuk admin/pengurus (SEC-101)
    all_parts = []
    par_cur = db.participants.find({'status': 'aktif'}, {'_id': 1, 'name': 1, 'code': 1, 'phone': 1})
    async for p in par_cur:
        p['id'] = p.pop('_id')
        all_parts.append(p)
    return {'items': items, 'all_participants': all_parts}


@api.get('/attendance/summary')
async def summary(range_key: Literal['daily', 'weekly'] = Query('daily'), date_f: Optional[str] = Query(None, alias='date'), user: dict = Depends(require_roles('admin', 'pengurus'))):
    today = datetime.now().date()
    if range_key == 'daily':
        d = date.fromisoformat(date_f) if date_f else today
        start, end = d, d
    else:
        d = date.fromisoformat(date_f) if date_f else today
        start = d - timedelta(days=d.weekday())
        end = start + timedelta(days=6)
    filt = {'activity_date': {'$gte': start.isoformat(), '$lte': end.isoformat()}}
    cur = db.attendance.find(filt)
    counts = {'hadir': 0, 'izin': 0, 'alpha': 0}
    by_activity_map: dict = {}
    async for a in cur:
        counts[a['status']] = counts.get(a['status'], 0) + 1
        aid = a['activity_id']
        by_activity_map.setdefault(aid, {'activity_name': a.get('activity_name', ''), 'activity_date': a.get('activity_date', ''), 'hadir': 0, 'izin': 0, 'alpha': 0})
        by_activity_map[aid][a['status']] += 1
    return {
        'range': range_key,
        'start': start.isoformat(),
        'end': end.isoformat(),
        'counts': counts,
        'by_activity': list(by_activity_map.values()),
    }


# --------- Reports (menu Laporan) — memakai data absensi yang sudah ada ---------
def _build_report_filter(date_from: Optional[str], date_to: Optional[str], activity_id: Optional[str], q: Optional[str], status_f: Optional[str]) -> dict:
    filt: dict = {}
    if date_from or date_to:
        rng: dict = {}
        if date_from:
            rng['$gte'] = date_from
        if date_to:
            rng['$lte'] = date_to
        filt['activity_date'] = rng
    if activity_id:
        filt['activity_id'] = activity_id
    if q:
        filt['participant_name'] = {'$regex': re.escape(q), '$options': 'i'}
    if status_f in ('hadir', 'izin', 'alpha'):
        filt['status'] = status_f
    return filt


@api.get('/reports/attendance')
async def report_attendance(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    activity_id: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    status_f: Optional[str] = Query(None, alias='status'),
    user: dict = Depends(require_roles('admin', 'pengurus')),
):
    filt = _build_report_filter(date_from, date_to, activity_id, q, status_f)
    counts = {'hadir': 0, 'izin': 0, 'alpha': 0}
    parts = set()
    rows = []
    async for a in db.attendance.find(filt).sort([('activity_date', -1), ('activity_name', 1), ('participant_name', 1)]).limit(5000):
        st = a.get('status', '')
        counts[st] = counts.get(st, 0) + 1
        if a.get('participant_id'):
            parts.add(a['participant_id'])
        rows.append({
            'activity_date': a.get('activity_date', ''),
            'activity_name': a.get('activity_name', ''),
            'participant_name': a.get('participant_name', ''),
            'participant_code': a.get('participant_code', ''),
            'status': st,
            'time_in': a.get('time_in', ''),
            'method': a.get('method', ''),
        })
    total = counts['hadir'] + counts['izin'] + counts['alpha']
    rate = round(counts['hadir'] / total * 100, 1) if total else 0
    return {
        'rows': rows,
        'stats': {
            'total_peserta': len(parts),
            'hadir': counts['hadir'],
            'izin': counts['izin'],
            'alpha': counts['alpha'],
            'total': total,
            'rate_hadir': rate,
        },
        'total': total,
    }


@api.get('/reports/attendance/export/xlsx')
async def report_attendance_xlsx(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    activity_id: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    status_f: Optional[str] = Query(None, alias='status'),
    user: dict = Depends(require_roles('admin', 'pengurus')),
):
    filt = _build_report_filter(date_from, date_to, activity_id, q, status_f)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Laporan Absensi'
    ws.append(['Tanggal Kegiatan', 'Kegiatan', 'Peserta', 'Kode', 'Status', 'Jam (WITA)', 'Metode'])
    counts = {'hadir': 0, 'izin': 0, 'alpha': 0}
    parts = set()
    async for r in db.attendance.find(filt).sort([('activity_date', -1), ('activity_name', 1), ('participant_name', 1)]).limit(20000):
        st = r.get('status', '')
        counts[st] = counts.get(st, 0) + 1
        if r.get('participant_id'):
            parts.add(r['participant_id'])
        ws.append([
            r.get('activity_date', ''), r.get('activity_name', ''), r.get('participant_name', ''),
            r.get('participant_code', ''), st, r.get('time_in', ''), r.get('method', ''),
        ])
    total = counts['hadir'] + counts['izin'] + counts['alpha']
    rate = round(counts['hadir'] / total * 100, 1) if total else 0
    ws2 = wb.create_sheet('Rekap')
    period = f"{date_from or '-'} s/d {date_to or '-'}" if (date_from or date_to) else 'Semua tanggal'
    ws2.append(['Periode', period])
    ws2.append(['Total Peserta', len(parts)])
    ws2.append(['Hadir', counts['hadir']])
    ws2.append(['Izin', counts['izin']])
    ws2.append(['Alpha', counts['alpha']])
    ws2.append(['Persentase Kehadiran', f'{rate}%'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Laporan_Absensi_{date_from or datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename="{fname}"'})


@api.get('/reports/monthly')
async def report_monthly(
    month: str = Query(..., description='Format YYYY-MM'),
    user: dict = Depends(require_roles('admin', 'pengurus')),
):
    if not re.fullmatch(r'\d{4}-\d{2}', month or ''):
        raise HTTPException(status_code=400, detail='Format bulan harus YYYY-MM')
    filt = {'activity_date': {'$gte': f'{month}-01', '$lte': f'{month}-31'}}
    by_activity: dict = {}
    counts = {'hadir': 0, 'izin': 0, 'alpha': 0}
    parts = set()
    async for a in db.attendance.find(filt):
        st = a.get('status', '')
        counts[st] = counts.get(st, 0) + 1
        if a.get('participant_id'):
            parts.add(a['participant_id'])
        aid = a.get('activity_id', '')
        row = by_activity.setdefault(aid, {
            'activity_id': aid,
            'activity_name': a.get('activity_name', ''),
            'activity_date': a.get('activity_date', ''),
            'hadir': 0, 'izin': 0, 'alpha': 0,
        })
        if st in row:
            row[st] += 1
    out_rows = []
    for r in by_activity.values():
        tot = r['hadir'] + r['izin'] + r['alpha']
        r['total'] = tot
        r['rate'] = round(r['hadir'] / tot * 100, 1) if tot else 0
        out_rows.append(r)
    out_rows.sort(key=lambda x: (x['activity_date'], x['activity_name']))
    total = counts['hadir'] + counts['izin'] + counts['alpha']
    return {
        'month': month,
        'by_activity': out_rows,
        'totals': {
            'activities': len(out_rows),
            'total_peserta': len(parts),
            'hadir': counts['hadir'],
            'izin': counts['izin'],
            'alpha': counts['alpha'],
            'total': total,
            'rate_hadir': round(counts['hadir'] / total * 100, 1) if total else 0,
        },
    }


# --------- Musyawarah ---------
@api.get('/musyawarah')
async def list_musyawarah(kind: Optional[str] = None, user: dict = Depends(require_roles('admin', 'pengurus'))):
    filt = {'kind': kind} if kind else {}
    cur = db.musyawarah.find(filt).sort('updated_at', -1)
    items = []
    async for m in cur:
        m['id'] = m.pop('_id')
        items.append(m)
    return {'items': items}


@api.post('/musyawarah')
async def create_musyawarah(body: MusyawarahCreate, user: dict = Depends(require_roles('admin', 'pengurus'))):
    doc = {
        '_id': new_id(),
        'kind': body.kind,
        'title': body.title,
        'content': body.content,
        'date': body.date or datetime.now().strftime('%Y-%m-%d'),
        'created_by': user['id'],
        'created_by_name': user.get('name', ''),
        'created_at': iso(now_utc()),
        'updated_at': iso(now_utc()),
    }
    await db.musyawarah.insert_one(doc)
    await log_activity(user['id'], user.get('name', ''), 'create_musyawarah', 'musyawarah', doc['_id'])
    doc['id'] = doc.pop('_id')
    return doc


@api.get('/musyawarah/{mid}')
async def get_musyawarah(mid: str, user: dict = Depends(require_roles('admin', 'pengurus'))):
    m = await db.musyawarah.find_one({'_id': mid})
    if not m:
        raise HTTPException(status_code=404, detail='Catatan tidak ditemukan')
    m['id'] = m.pop('_id')
    return m


@api.patch('/musyawarah/{mid}')
async def update_musyawarah(mid: str, body: MusyawarahUpdate, user: dict = Depends(require_roles('admin', 'pengurus'))):
    upd = {k: v for k, v in body.model_dump().items() if v is not None}
    upd['updated_at'] = iso(now_utc())
    r = await db.musyawarah.update_one({'_id': mid}, {'$set': upd})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail='Catatan tidak ditemukan')
    return {'ok': True, 'updated_at': upd['updated_at']}


@api.delete('/musyawarah/{mid}')
async def delete_musyawarah(mid: str, user: dict = Depends(require_roles('admin'))):
    r = await db.musyawarah.delete_one({'_id': mid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Catatan tidak ditemukan')
    return {'ok': True}


@api.get('/musyawarah/{mid}/export/xlsx')
async def export_musyawarah_xlsx(mid: str, user: dict = Depends(require_roles('admin', 'pengurus'))):
    m = await db.musyawarah.find_one({'_id': mid})
    if not m:
        raise HTTPException(status_code=404, detail='Catatan tidak ditemukan')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Catatan'
    ws.append(['E-Kertalangu | Catatan Musyawarah'])
    ws.append(['Jenis', m.get('kind', '')])
    ws.append(['Judul', m.get('title', '')])
    ws.append(['Tanggal', m.get('date', '')])
    ws.append(['Dibuat oleh', m.get('created_by_name', '')])
    ws.append([])
    ws.append(['Isi'])
    for line in (m.get('content') or '').splitlines():
        ws.append([line])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'Musyawarah_{m.get("kind","")}_{m.get("date","")}.xlsx'
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename="{fname}"'})


# --------- Activity log ---------
@api.get('/activity-log')
async def get_logs(limit: int = 200, user: dict = Depends(require_roles('admin'))):
    cur = db.activity_logs.find({}).sort('timestamp', -1).limit(limit)
    items = []
    async for l in cur:
        l['id'] = l.pop('_id')
        items.append(l)
    return {'items': items}


# --------- WhatsApp templates ---------
@api.get('/wa/templates')
async def wa_templates(user: dict = Depends(get_current_user)):
    today = datetime.now().strftime('%d %B %Y')
    return {'templates': [
        {'id': 'harian', 'title': 'Rekap Harian', 'body': f'Assalamu\'alaikum. Berikut rekap kehadiran harian E-Kertalangu tanggal {today}. Jazakumullah khoiro.'},
        {'id': 'mingguan', 'title': 'Rekap Mingguan', 'body': f'Assalamu\'alaikum. Terlampir rekap kehadiran pekan ini per {today}. Barakallah.'},
        {'id': 'musyawarah', 'title': 'Catatan Musyawarah', 'body': f'Assalamu\'alaikum. Berikut catatan musyawarah (4S / Tim 7) tanggal {today}.'},
        {'id': 'reminder', 'title': 'Pengingat Kehadiran', 'body': 'Assalamu\'alaikum. Sekedar mengingatkan, kegiatan pengajian akan berlangsung. Mohon kehadirannya. Jazakumullah khoiro.'},
    ]}


# --------- Backup / Full Export ---------
@api.get('/backup/xlsx')
async def backup_xlsx(user: dict = Depends(require_roles('admin', 'pengurus'))):
    wb = Workbook()
    # Participants
    ws = wb.active
    ws.title = 'Peserta'
    ws.append(['Kode', 'Nama', 'Gender', 'TTL', 'HP', 'Email', 'Pendidikan', 'Status', 'Tag Rahasia'])
    async for p in db.participants.find({}).sort('code', 1):
        ws.append([
            p.get('code', ''), p.get('name', ''), p.get('gender', ''),
            f"{p.get('birth_place','')} / {p.get('birth_date','')}",
            p.get('phone', ''), p.get('email', ''), p.get('education', ''),
            p.get('status', ''), 'Ya' if p.get('is_secret_tag') else '',
        ])
    # Activities
    ws2 = wb.create_sheet('Kegiatan')
    ws2.append(['Kode', 'Nama', 'Jenis', 'Tanggal', 'Mulai (WITA)', 'Selesai (WITA)', 'Lokasi', 'GPS Lat', 'GPS Lng', 'Di Luar', 'Rahasia'])
    async for a in db.activities.find({}).sort('date', -1):
        ws2.append([
            a.get('code', ''), a.get('name', ''), a.get('type', ''), a.get('date', ''),
            a.get('start_time', ''), a.get('end_time', ''), a.get('location', ''),
            a.get('gps_lat'), a.get('gps_lng'),
            'Ya' if a.get('is_outside') else '', 'Ya' if a.get('is_secret') else '',
        ])
    # Attendance
    ws3 = wb.create_sheet('Absensi')
    ws3.append(['Tanggal Kegiatan', 'Kegiatan', 'Peserta', 'Kode', 'Status', 'Jam (WITA)', 'Metode', 'Dicatat oleh', 'Tercatat'])
    async for r in db.attendance.find({}).sort('recorded_at', -1):
        ws3.append([
            r.get('activity_date', ''), r.get('activity_name', ''), r.get('participant_name', ''), r.get('participant_code', ''),
            r.get('status', ''), r.get('time_in', ''), r.get('method', ''), r.get('recorded_by_name', ''), r.get('recorded_at', ''),
        ])
    # Musyawarah
    ws4 = wb.create_sheet('Musyawarah')
    ws4.append(['Jenis', 'Judul', 'Tanggal', 'Isi', 'Dibuat oleh', 'Diupdate'])
    async for m in db.musyawarah.find({}).sort('date', -1):
        ws4.append([m.get('kind', ''), m.get('title', ''), m.get('date', ''), (m.get('content', '') or '')[:32000], m.get('created_by_name', ''), m.get('updated_at', '')])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = datetime.now().strftime('%d-%m-%Y')
    fname = f'e kertalangu {today}.xlsx'
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename="{fname}"'})


# --------- Dashboard KPIs ---------
# --------- Self-check-in (peserta) ---------
class SelfCheckin(BaseModel):
    activity_qr: str


@api.post('/attendance/self')
async def attendance_self(body: SelfCheckin, user: dict = Depends(get_current_user)):
    par = await _resolve_self_participant(user)
    if not par:
        if not user.get('participant_id'):
            raise HTTPException(status_code=400, detail='Akun peserta belum tertaut ke data peserta. Silakan hubungi Admin.')
        raise HTTPException(status_code=404, detail='Data peserta tidak ditemukan. Silakan hubungi Admin.')
    payload = (body.activity_qr or '').strip()
    if not payload.startswith('EKTL:A:'):
        raise HTTPException(status_code=400, detail='QR bukan QR kegiatan')
    code = payload.split('EKTL:A:', 1)[1]
    act = await db.activities.find_one({'code': code})
    if not act:
        raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan atau sudah tidak tersedia.')
    # Rule: only allow check-in on the day of activity
    today = datetime.now().strftime('%Y-%m-%d')
    if act.get('date') != today:
        raise HTTPException(status_code=400, detail=f'Kegiatan ini bukan hari ini ({act.get("date")}). Absen mandiri hanya di hari kegiatan.')
    return await _record_attendance(act['_id'], par['_id'], 'hadir', None, 'self', 'Absen mandiri (self scan)', user)


# --------- Public QR untuk pendaftaran ---------
@api.get('/qr/register-public')
async def qr_register_public(base_url: Optional[str] = None):
    # Return QR image that opens /register on the frontend
    url = (base_url or '').rstrip('/') + '/register' if base_url else '/register'
    png = qr_png_bytes(url)
    return StreamingResponse(io.BytesIO(png), media_type='image/png')


# --------- Reminders (tomorrow) ---------
@api.get('/reminders/tomorrow')
async def reminders_tomorrow(user: dict = Depends(require_roles('admin', 'pengurus'))):
    tomorrow = (datetime.now().date() + timedelta(days=1)).isoformat()
    acts = []
    async for a in db.activities.find({'date': tomorrow}).sort('start_time', 1):
        a['id'] = a.pop('_id')
        acts.append(a)
    # active peserta with phones
    parts = []
    async for p in db.participants.find({'status': 'aktif', 'phone': {'$ne': ''}}):
        parts.append({'id': p['_id'], 'name': p['name'], 'phone': p.get('phone', '')})
    return {'date': tomorrow, 'activities': acts, 'participants': parts}


# --------- Dashboard extended stats ---------
@api.get('/dashboard/stats')
async def dashboard_stats(user: dict = Depends(get_current_user)):
    today = datetime.now().strftime('%Y-%m-%d')
    # gender counts (aktif only)
    male = await db.participants.count_documents({'status': 'aktif', 'gender': 'L'})
    female = await db.participants.count_documents({'status': 'aktif', 'gender': 'P'})
    # attendance last 30 days
    d30 = (datetime.now().date() - timedelta(days=29)).isoformat()
    pipe = [
        {'$match': {'activity_date': {'$gte': d30}}},
        {'$group': {'_id': '$status', 'n': {'$sum': 1}}},
    ]
    counts = {'hadir': 0, 'izin': 0, 'alpha': 0}
    async for r in db.attendance.aggregate(pipe):
        counts[r['_id']] = r['n']
    total = counts['hadir'] + counts['izin'] + counts['alpha']
    rate = round((counts['hadir'] / total) * 100, 1) if total else 0.0
    # today counts by status
    today_pipe = [
        {'$match': {'activity_date': today}},
        {'$group': {'_id': '$status', 'n': {'$sum': 1}}},
    ]
    today_counts = {'hadir': 0, 'izin': 0, 'alpha': 0}
    async for r in db.attendance.aggregate(today_pipe):
        today_counts[r['_id']] = r['n']
    # per-day series last 14 days
    d14 = (datetime.now().date() - timedelta(days=13)).isoformat()
    series_pipe = [
        {'$match': {'activity_date': {'$gte': d14}, 'status': 'hadir'}},
        {'$group': {'_id': '$activity_date', 'n': {'$sum': 1}}},
        {'$sort': {'_id': 1}},
    ]
    series = []
    async for r in db.attendance.aggregate(series_pipe):
        series.append({'date': r['_id'], 'hadir': r['n']})
    # pending users
    pending = await db.users.count_documents({'pending_approval': True, 'active': False})
    result = {
        'gender': {'L': male, 'P': female, 'total': male + female},
        'last30': {'counts': counts, 'total': total, 'rate_hadir': rate},
        'today': today_counts,
        'series_hadir_14d': series,
        'pending_users': pending,
    }
    # Role-based aggregate stats (admin/pengurus). No PII, counts only.
    roles = user.get('roles') or [user.get('role')]
    if 'admin' in roles or 'pengurus' in roles:
        total_participants = await db.participants.count_documents({'status': 'aktif'})
        linked_ids = await db.users.distinct('participant_id', {'participant_id': {'$ne': None}})
        activated = await db.participants.count_documents({'status': 'aktif', '_id': {'$in': linked_ids}})
        unactivated = max(total_participants - activated, 0)
        total_activities = await db.activities.count_documents({})
        month_prefix = datetime.now().strftime('%Y-%m')
        monthly_activities = await db.activities.count_documents({'date': {'$regex': f'^{month_prefix}'}})
        total_attendance = await db.attendance.count_documents({})
        att_hadir_all = await db.attendance.count_documents({'status': 'hadir'})
        att_rate_all = round((att_hadir_all / total_attendance) * 100, 1) if total_attendance else 0.0
        result['overview'] = {
            'total_participants': total_participants,
            'activated_participants': activated,
            'unactivated_participants': unactivated,
            'total_activities': total_activities,
            'monthly_activities': monthly_activities,
            'total_attendance': total_attendance,
            'attendance_rate': att_rate_all,
        }
    return result


@api.get('/dashboard/kpi')
async def dashboard_kpi(user: dict = Depends(get_current_user)):
    today = datetime.now().strftime('%Y-%m-%d')
    total_p = await db.participants.count_documents({'status': 'aktif'})
    total_p_all = await db.participants.count_documents({})
    total_a_upcoming = await db.activities.count_documents({'date': {'$gte': today}})
    total_a = await db.activities.count_documents({})
    hadir_today = await db.attendance.count_documents({'activity_date': today, 'status': 'hadir'})
    return {
        'peserta_aktif': total_p,
        'peserta_total': total_p_all,
        'kegiatan_upcoming': total_a_upcoming,
        'kegiatan_total': total_a,
        'hadir_hari_ini': hadir_today,
        'today': today,
    }


# --------- Rotating QR & GPS self check-in helpers ---------
ROTATE_WINDOW_SEC = 30


def rotating_token(activity_code: str, window: int) -> str:
    msg = f"{activity_code}:{window}"
    return hmac.new(JWT_SECRET.encode(), msg.encode(), hashlib.sha256).hexdigest()[:12]


def current_window() -> int:
    return int(time.time() // ROTATE_WINDOW_SEC)


def haversine_m(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    R = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


@api.get('/activities/{aid}/qr-current')
async def get_qr_current(aid: str, user: dict = Depends(get_current_user)):
    a = await db.activities.find_one({'_id': aid})
    if not a:
        raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan')
    win = current_window()
    tok = rotating_token(a['code'], win)
    payload = f"EKTL:AR:{a['code']}:{win}:{tok}"
    now_s = int(time.time())
    remain = ROTATE_WINDOW_SEC - (now_s % ROTATE_WINDOW_SEC)
    return {
        'payload': payload,
        'datauri': qr_png_datauri(payload),
        'window': win,
        'expires_in': remain,
        'window_sec': ROTATE_WINDOW_SEC,
    }


class SelfCheckinV2(BaseModel):
    activity_qr: str
    lat: Optional[float] = None
    lng: Optional[float] = None


@api.post('/attendance/self-v2')
async def attendance_self_v2(body: SelfCheckinV2, user: dict = Depends(get_current_user)):
    par = await _resolve_self_participant(user)
    if not par:
        if not user.get('participant_id'):
            raise HTTPException(status_code=400, detail='Akun peserta belum tertaut ke data peserta. Silakan hubungi Admin.')
        raise HTTPException(status_code=404, detail='Data peserta tidak ditemukan. Silakan hubungi Admin.')
    if par.get('status') == 'arsip':
        raise HTTPException(status_code=400, detail='Akun peserta telah diarsipkan, tidak dapat melakukan absensi. Hubungi Admin.')
    payload = (body.activity_qr or '').strip()

    act = None
    # QR format URL (…/a/<code>) — normalisasi ke code (defense-in-depth)
    if not payload.startswith('EKTL:') and '/a/' in payload:
        m = re.search(r'/a/([A-Za-z0-9_-]+)', payload)
        if m:
            payload = f'EKTL:A:{m.group(1)}'
    # Rotating QR: EKTL:AR:{code}:{window}:{token}
    if payload.startswith('EKTL:AR:'):
        parts = payload.split(':')
        if len(parts) < 5:
            raise HTTPException(status_code=400, detail='QR kegiatan tidak valid.')
        code, win_str, tok = parts[2], parts[3], parts[4]
        try:
            win = int(win_str)
        except ValueError:
            raise HTTPException(status_code=400, detail='QR kegiatan tidak valid.')
        cur = current_window()
        if abs(cur - win) > 1:
            raise HTTPException(status_code=400, detail='QR kegiatan sudah expired. Minta QR baru dari pengurus.')
        expected = rotating_token(code, win)
        if not hmac.compare_digest(expected, tok):
            raise HTTPException(status_code=400, detail='QR kegiatan tidak valid (token mismatch).')
        act = await db.activities.find_one({'code': code})
    elif payload.startswith('EKTL:A:'):
        code = payload.split('EKTL:A:', 1)[1]
        act = await db.activities.find_one({'code': code})
    else:
        raise HTTPException(status_code=400, detail='QR kegiatan tidak valid.')

    if not act:
        raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan atau sudah tidak tersedia.')

    today = datetime.now().strftime('%Y-%m-%d')
    if act.get('date') != today:
        raise HTTPException(status_code=400, detail=f"Kegiatan bukan hari ini ({act.get('date')}). Absen mandiri hanya di hari kegiatan.")

    # GPS radius check (if activity has coordinates and radius)
    if act.get('gps_lat') is not None and act.get('gps_lng') is not None:
        if body.lat is None or body.lng is None:
            raise HTTPException(status_code=400, detail='Aktifkan izin lokasi (GPS) untuk absen mandiri.')
        dist = haversine_m(act['gps_lat'], act['gps_lng'], body.lat, body.lng)
        radius = int(act.get('radius_m') or 100)
        if dist > radius:
            raise HTTPException(status_code=400, detail=f'Anda berada {int(dist)}m dari lokasi. Radius maks {radius}m.')

    # Anti-double: jika sudah hadir di kegiatan ini, beri pesan jelas (tanpa membuat data ganda)
    existing = await db.attendance.find_one({'activity_id': act['_id'], 'participant_id': par['_id']})
    if existing and existing.get('status') == 'hadir':
        raise HTTPException(status_code=400, detail='Anda sudah melakukan absensi pada kegiatan ini.')

    return await _record_attendance(act['_id'], par['_id'], 'hadir', None, 'self', 'Absen mandiri', user)


# --------- Recurring activities ---------
class RecurringReq(BaseModel):
    base: ActivityCreate
    weeks: int = 4  # jumlah minggu berulang


@api.post('/activities/recurring')
async def create_recurring(body: RecurringReq, user: dict = Depends(require_roles('admin', 'pengurus'))):
    if body.weeks < 1 or body.weeks > 52:
        raise HTTPException(status_code=400, detail='Weeks harus 1-52')
    try:
        base_date = date.fromisoformat(body.base.date)
    except Exception:
        raise HTTPException(status_code=400, detail='Format tanggal salah (YYYY-MM-DD)')
    created = []
    for w in range(body.weeks):
        d = (base_date + timedelta(weeks=w)).isoformat()
        code = await _next_activity_code()
        qr_payload = f'EKTL:A:{code}'
        doc = {
            '_id': new_id(),
            'code': code,
            'name': body.base.name,
            'type': body.base.type,
            'date': d,
            'start_time': body.base.start_time,
            'end_time': body.base.end_time,
            'location': body.base.location or 'Kertalangu',
            'gps_lat': body.base.gps_lat,
            'gps_lng': body.base.gps_lng,
            'radius_m': int(body.base.radius_m or 100),
            'is_outside': bool(body.base.is_outside),
            'is_secret': bool(body.base.is_secret),
            'secret_allow': [],
            'pengajar': body.base.pengajar or '',
            'materi_progress': body.base.materi_progress or '',
            'notes': body.base.notes or '',
            'qr_payload': qr_payload,
            'created_at': iso(now_utc()),
            'created_by': user['id'],
            'recurring_group': body.base.name,
        }
        await db.activities.insert_one(doc)
        created.append(code)
    await log_activity(user['id'], user.get('name', ''), 'create_recurring', 'activity', '', {'name': body.base.name, 'weeks': body.weeks, 'count': len(created)})
    return {'ok': True, 'count': len(created), 'codes': created}


# --------- Announcements ---------
class AnnouncementCreate(BaseModel):
    title: str
    body: str = ''
    pengajar: Optional[str] = ''
    materi_progress: Optional[str] = ''
    activity_id: Optional[str] = None
    priority: Literal['normal', 'penting'] = 'normal'
    pinned: bool = False


class AnnouncementUpdate(BaseModel):
    title: Optional[str] = None
    body: Optional[str] = None
    pengajar: Optional[str] = None
    materi_progress: Optional[str] = None
    activity_id: Optional[str] = None
    priority: Optional[Literal['normal', 'penting']] = None
    pinned: Optional[bool] = None


@api.get('/announcements')
async def list_announcements(user: dict = Depends(get_current_user)):
    cur = db.announcements.find({}).sort([('pinned', -1), ('created_at', -1)]).limit(100)
    items = []
    async for a in cur:
        a['id'] = a.pop('_id')
        items.append(a)
    return {'items': items}


@api.post('/announcements')
async def create_announcement(body: AnnouncementCreate, user: dict = Depends(require_roles('admin', 'pengurus'))):
    doc = {
        '_id': new_id(),
        'title': body.title,
        'body': body.body,
        'pengajar': body.pengajar or '',
        'materi_progress': body.materi_progress or '',
        'activity_id': body.activity_id,
        'priority': body.priority,
        'pinned': bool(body.pinned),
        'created_at': iso(now_utc()),
        'created_by': user['id'],
        'created_by_name': user.get('name', ''),
    }
    await db.announcements.insert_one(doc)
    await log_activity(user['id'], user.get('name', ''), 'create_announcement', 'announcement', doc['_id'], {'title': body.title})
    doc['id'] = doc.pop('_id')
    return doc


@api.patch('/announcements/{aid}')
async def update_announcement(aid: str, body: AnnouncementUpdate, user: dict = Depends(require_roles('admin', 'pengurus'))):
    upd = {k: v for k, v in body.model_dump().items() if v is not None}
    upd['updated_at'] = iso(now_utc())
    r = await db.announcements.update_one({'_id': aid}, {'$set': upd})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail='Pengumuman tidak ditemukan')
    return {'ok': True}


@api.delete('/announcements/{aid}')
async def delete_announcement(aid: str, user: dict = Depends(require_roles('admin', 'pengurus'))):
    r = await db.announcements.delete_one({'_id': aid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Pengumuman tidak ditemukan')
    return {'ok': True}


# --------- Alpha Alert ---------
@api.get('/analytics/alpha-alert')
async def alpha_alert(threshold: int = 3, user: dict = Depends(require_roles('admin', 'pengurus'))):
    """Peserta dengan >=threshold alpha dalam 30 hari terakhir."""
    d30 = (datetime.now().date() - timedelta(days=29)).isoformat()
    pipe = [
        {'$match': {'activity_date': {'$gte': d30}, 'status': 'alpha'}},
        {'$group': {'_id': '$participant_id', 'n': {'$sum': 1}, 'name': {'$first': '$participant_name'}, 'code': {'$first': '$participant_code'}}},
        {'$match': {'n': {'$gte': threshold}}},
        {'$sort': {'n': -1}},
    ]
    results = [r async for r in db.attendance.aggregate(pipe)]
    pids = [r['_id'] for r in results]
    phones = {}
    if pids:
        async for p in db.participants.find({'_id': {'$in': pids}}, {'phone': 1}):
            phones[p['_id']] = p.get('phone', '')
    items = []
    for r in results:
        items.append({
            'participant_id': r['_id'],
            'name': r.get('name', ''),
            'code': r.get('code', ''),
            'phone': phones.get(r['_id'], ''),
            'alpha_count': r['n'],
        })
    return {'items': items, 'threshold': threshold, 'range_start': d30}


# --------- XLSX import Peserta ---------
@api.post('/participants/import-xlsx')
async def import_participants_xlsx(file: UploadFile = File(...), user: dict = Depends(require_roles('admin', 'pengurus'))):
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='File harus .xlsx')
    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents))
    except Exception:
        raise HTTPException(status_code=400, detail='File .xlsx tidak dapat dibaca')
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'ok': True, 'results': [], 'count': 0}
    # Header row detection — assume first row is header if any cell contains "nama"
    header_used = False
    if rows and any(isinstance(c, str) and 'nama' in c.lower() for c in rows[0] if c):
        header_used = True
        data_rows = rows[1:]
    else:
        data_rows = rows

    results = []
    for r in data_rows:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        gender = 'L'
        if len(r) > 1 and r[1]:
            gender = 'P' if str(r[1]).strip().upper().startswith('P') else 'L'
        phone = str(r[2]).strip() if len(r) > 2 and r[2] else ''
        email = str(r[3]).strip() if len(r) > 3 and r[3] else ''
        birth_place = str(r[4]).strip() if len(r) > 4 and r[4] else ''
        birth_date = str(r[5]).strip() if len(r) > 5 and r[5] else ''
        education = str(r[6]).strip() if len(r) > 6 and r[6] else ''
        try:
            pc = ParticipantCreate(
                name=name, gender=gender, phone=phone, email=email,
                birth_place=birth_place, birth_date=birth_date, education=education,
                duplicate_action='append',
            )
            r_ok = await create_participant(pc, user)  # type: ignore
            results.append({'ok': True, 'name': r_ok['name'], 'code': r_ok['code']})
        except HTTPException as e:
            results.append({'ok': False, 'name': name, 'error': str(e.detail)})
        except Exception as e:
            results.append({'ok': False, 'name': name, 'error': str(e)})
    ok = sum(1 for x in results if x['ok'])
    await log_activity(user['id'], user.get('name', ''), 'import_participants', 'participant', '', {'count': ok, 'total': len(results)})
    return {'ok': True, 'count': ok, 'total': len(results), 'results': results, 'header_used': header_used}


# --------- Photo Album per Activity (disimpan di MongoDB agar aman saat deploy) ---------


@api.post('/activities/{aid}/photos')
async def upload_photo(aid: str, file: UploadFile = File(...), caption: str = Form(''), user: dict = Depends(require_roles('admin', 'pengurus'))):
    act = await db.activities.find_one({'_id': aid})
    if not act:
        raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan')
    if not file.content_type or not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail='File harus berupa gambar')
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail='Ukuran gambar maksimal 5MB')
    ext = (file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'jpg').lower()
    if ext not in ('jpg', 'jpeg', 'png', 'webp'):
        ext = 'jpg'
    photo_id = new_id()
    doc = {
        '_id': photo_id,
        'activity_id': aid,
        'filename': f'{photo_id}.{ext}',
        'content_type': file.content_type,
        'data_b64': base64.b64encode(contents).decode('ascii'),
        'caption': caption,
        'uploaded_by': user['id'],
        'uploaded_by_name': user.get('name', ''),
        'uploaded_at': iso(now_utc()),
        'url': f'/api/uploads/photos/{aid}/{photo_id}.{ext}',
    }
    await db.photos.insert_one(doc)
    return {k: v for k, v in doc.items() if k not in ('data_b64', '_id')} | {'id': photo_id}


@api.get('/uploads/photos/{aid}/{filename}')
async def serve_photo(aid: str, filename: str):
    p = await db.photos.find_one({'activity_id': aid, 'filename': filename})
    if not p or not p.get('data_b64'):
        raise HTTPException(status_code=404, detail='Foto tidak ditemukan')
    data = base64.b64decode(p['data_b64'])
    return StreamingResponse(io.BytesIO(data), media_type=p.get('content_type', 'image/jpeg'),
                             headers={'Cache-Control': 'public, max-age=31536000'})


@api.get('/activities/{aid}/photos')
async def list_photos(aid: str, user: dict = Depends(get_current_user)):
    cur = db.photos.find({'activity_id': aid}, {'data_b64': 0}).sort('uploaded_at', -1)
    items = []
    async for p in cur:
        p['id'] = p.pop('_id')
        items.append(p)
    return {'items': items}


@api.delete('/photos/{pid}')
async def delete_photo(pid: str, user: dict = Depends(require_roles('admin', 'pengurus'))):
    p = await db.photos.find_one({'_id': pid})
    if not p:
        raise HTTPException(status_code=404, detail='Foto tidak ditemukan')
    await db.photos.delete_one({'_id': pid})
    return {'ok': True}


# --------- Password reset (admin-driven, tanpa simpan password teks-polos) ---------
RESET_TTL_MIN = 60


def _gen_temp_password() -> str:
    # 10 karakter acak, mudah dibaca, tanpa karakter ambigu
    alphabet = 'ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnpqrstuvwxyz23456789'
    return ''.join(secrets.choice(alphabet) for _ in range(10))


class ForgotPwReq(BaseModel):
    identifier: str


@api.post('/auth/forgot-password')
async def forgot_password(body: ForgotPwReq):
    """Buat permintaan reset. TIDAK menyimpan password. Admin akan membuat password sementara."""
    ident = (body.identifier or '').strip()
    if len(ident) < 3:
        raise HTTPException(status_code=400, detail='Identitas wajib diisi')
    q = {'$or': [{'email': ident.lower()}, {'username': ident.lower()}, {'phone': normalize_phone(ident) or ident}]}
    user = await db.users.find_one(q)
    doc = {
        '_id': new_id(),
        'identifier': ident,
        'user_id': user['_id'] if user else None,
        'user_name': user.get('name') if user else '',
        'created_at': iso(now_utc()),
        'expires_at': iso(now_utc() + timedelta(minutes=RESET_TTL_MIN)),
        'status': 'pending',
    }
    await db.password_resets.insert_one(doc)
    # Pesan seragam (tidak membocorkan apakah akun ada)
    return {'ok': True, 'message': 'Permintaan reset diterima. Admin akan membuatkan password sementara dan menghubungi Anda.'}


@api.get('/auth/reset-requests')
async def list_reset_requests(user: dict = Depends(require_roles('admin'))):
    cur = db.password_resets.find({'status': 'pending'}).sort('created_at', -1)
    items = []
    async for r in cur:
        r['id'] = r.pop('_id')
        r.pop('requested_new_password', None)  # buang field lama bila ada
        items.append(r)
    return {'items': items}


@api.post('/auth/reset-requests/{rid}/apply')
async def apply_reset(rid: str, user: dict = Depends(require_roles('admin'))):
    r = await db.password_resets.find_one({'_id': rid})
    if not r:
        raise HTTPException(status_code=404, detail='Permintaan tidak ditemukan')
    if r.get('status') != 'pending':
        raise HTTPException(status_code=400, detail='Permintaan sudah diproses')
    if not r.get('user_id'):
        raise HTTPException(status_code=400, detail='User tidak ditemukan untuk permintaan ini')
    try:
        exp = datetime.fromisoformat(r['expires_at']) if r.get('expires_at') else None
    except Exception:
        exp = None
    if exp and now_utc() > exp:
        await db.password_resets.update_one({'_id': rid}, {'$set': {'status': 'expired'}})
        raise HTTPException(status_code=400, detail='Permintaan sudah kedaluwarsa. Minta peserta mengajukan ulang.')
    temp_pw = _gen_temp_password()
    target = await db.users.find_one({'_id': r['user_id']}, {'token_version': 1})
    new_tv = int((target or {}).get('token_version', 0)) + 1
    await db.users.update_one({'_id': r['user_id']}, {'$set': {'password_hash': hash_pw(temp_pw), 'token_version': new_tv}})
    await db.password_resets.update_one({'_id': rid}, {'$set': {'status': 'applied', 'applied_at': iso(now_utc()), 'applied_by': user['id']}})
    await log_activity(user['id'], user.get('name', ''), 'password_reset', 'user', r['user_id'])
    # Password sementara HANYA dikembalikan sekali ke admin untuk diteruskan; tidak disimpan.
    return {'ok': True, 'temp_password': temp_pw, 'user_name': r.get('user_name', '')}


@api.post('/auth/reset-requests/{rid}/reject')
async def reject_reset(rid: str, user: dict = Depends(require_roles('admin'))):
    r = await db.password_resets.update_one({'_id': rid}, {'$set': {'status': 'rejected', 'applied_at': iso(now_utc())}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail='Permintaan tidak ditemukan')
    return {'ok': True}


class AdminSetPwReq(BaseModel):
    new_password: str


@api.post('/users/{uid}/set-password')
async def admin_set_password(uid: str, body: AdminSetPwReq, user: dict = Depends(require_roles('admin'))):
    if len(body.new_password) < 6:
        raise HTTPException(status_code=400, detail='Password minimal 6 karakter')
    target = await db.users.find_one({'_id': uid}, {'token_version': 1})
    if not target:
        raise HTTPException(status_code=404, detail='Pengguna tidak ditemukan')
    new_tv = int(target.get('token_version', 0)) + 1
    await db.users.update_one({'_id': uid}, {'$set': {'password_hash': hash_pw(body.new_password), 'token_version': new_tv}})
    await log_activity(user['id'], user.get('name', ''), 'admin_set_password', 'user', uid)
    return {'ok': True}


# --------- Public Share (rekap kehadiran) ---------
class ShareCreate(BaseModel):
    kind: Literal['daily', 'weekly', 'monthly', 'activity'] = 'daily'
    date: Optional[str] = None
    activity_id: Optional[str] = None
    ttl_hours: int = 720  # 30 hari


@api.post('/share/attendance')
async def create_share(body: ShareCreate, user: dict = Depends(require_roles('admin', 'pengurus'))):
    token = secrets.token_urlsafe(12)
    doc = {
        '_id': token,
        'kind': body.kind,
        'date': body.date or datetime.now().strftime('%Y-%m-%d'),
        'activity_id': body.activity_id,
        'created_by': user['id'],
        'created_by_name': user.get('name', ''),
        'created_at': iso(now_utc()),
        'expires_at': iso(now_utc() + timedelta(hours=max(1, body.ttl_hours))),
    }
    await db.shares.insert_one(doc)
    return {'token': token, 'expires_at': doc['expires_at']}


@api.get('/share/attendance/{token}')
async def get_share(token: str):
    s = await db.shares.find_one({'_id': token})
    if not s:
        raise HTTPException(status_code=404, detail='Link tidak ditemukan')
    if datetime.fromisoformat(s['expires_at'].replace('Z', '+00:00')) < now_utc():
        raise HTTPException(status_code=410, detail='Link kadaluarsa')

    kind = s['kind']
    payload = {'kind': kind, 'date': s['date'], 'created_by_name': s.get('created_by_name', '')}

    if kind == 'activity' and s.get('activity_id'):
        act = await db.activities.find_one({'_id': s['activity_id']})
        if not act:
            raise HTTPException(status_code=404, detail='Kegiatan tidak ditemukan')
        rows = []
        counts = {'hadir': 0, 'izin': 0, 'alpha': 0}
        async for a in db.attendance.find({'activity_id': s['activity_id']}).sort('time_in', 1):
            rows.append({'name': a['participant_name'], 'status': a['status'], 'time_in': a.get('time_in', '')})
            counts[a['status']] = counts.get(a['status'], 0) + 1
        payload.update({
            'activity_name': act.get('name'), 'activity_date': act.get('date'),
            'pengajar': act.get('pengajar', ''), 'materi_progress': act.get('materi_progress', ''),
            'start_time': act.get('start_time'), 'end_time': act.get('end_time'),
            'counts': counts, 'rows': rows,
        })
    else:
        d = s['date']
        try:
            dd = date.fromisoformat(d)
        except Exception:
            dd = datetime.now().date()
        if kind == 'weekly':
            start = dd - timedelta(days=dd.weekday())
            end = start + timedelta(days=6)
            date_filter = {'$gte': start.isoformat(), '$lte': end.isoformat()}
            payload.update({'range_start': start.isoformat(), 'range_end': end.isoformat()})
        elif kind == 'monthly':
            start = dd.replace(day=1)
            nxt = start.replace(year=start.year + 1, month=1) if start.month == 12 else start.replace(month=start.month + 1)
            end = nxt - timedelta(days=1)
            date_filter = {'$gte': start.isoformat(), '$lte': end.isoformat()}
            payload.update({'range_start': start.isoformat(), 'range_end': end.isoformat()})
        else:
            date_filter = d
        counts = {'hadir': 0, 'izin': 0, 'alpha': 0}
        by_act: dict = {}
        rows = []
        cur = db.attendance.find({'activity_date': date_filter}).sort('activity_date', 1)
        async for a in cur:
            st = a['status']
            counts[st] = counts.get(st, 0) + 1
            aid = a['activity_id']
            by_act.setdefault(aid, {'name': a.get('activity_name', ''), 'date': a.get('activity_date', ''), 'hadir': 0, 'izin': 0, 'alpha': 0})
            by_act[aid][st] += 1
            rows.append({
                'name': a.get('participant_name', ''),
                'status': st,
                'activity_name': a.get('activity_name', ''),
                'time_in': a.get('time_in', ''),
            })
        payload.update({'counts': counts, 'activities': list(by_act.values()), 'rows': rows})
    return payload


app.include_router(api)

_cors_env = os.environ.get('CORS_ORIGINS', '*').strip()
_cors_origins = [o.strip() for o in _cors_env.split(',') if o.strip()]
_cors_wildcard = _cors_origins == ['*']
app.add_middleware(
    CORSMiddleware,
    allow_credentials=not _cors_wildcard,
    allow_origins=_cors_origins,
    allow_methods=['*'],
    allow_headers=['*'],
)


@app.get('/api/health')
async def health():
    return {'status': 'ok'}


# --------- Bootstrap: indexes + seed admin ---------
# Dijalankan sekali per proses. Di serverless (Vercel) event startup tidak selalu
# terpanggil, jadi bootstrap juga dipicu lazily pada request pertama.
_bootstrapped = False


async def bootstrap():
    global _bootstrapped
    if _bootstrapped:
        return
    _bootstrapped = True
    try:
        await db.users.create_index('email', sparse=True)
        await db.users.create_index('username', sparse=True)
        await db.users.create_index('phone', sparse=True)
        await db.participants.create_index('code', unique=True, sparse=True)
        await db.participants.create_index('name_key')
        await db.activities.create_index([('date', -1)])
        await db.attendance.create_index([('activity_id', 1), ('participant_id', 1)])
        await db.activity_logs.create_index([('timestamp', -1)])
    except Exception as e:
        logger.warning(f'Index create warning: {e}')

    # Seed admin (idempotent)
    admin_email = os.environ.get('ADMIN_EMAIL', 'admin@ekertalangu.local').lower()
    admin_pw = os.environ.get('ADMIN_PASSWORD', 'admin123')
    admin_name = os.environ.get('ADMIN_NAME', 'Admin')
    admin_username = os.environ.get('ADMIN_USERNAME', 'admin').lower()
    admin_phone = normalize_phone(os.environ.get('ADMIN_PHONE', ''))
    existing = await db.users.find_one({'email': admin_email})
    if existing is None:
        await db.users.insert_one({
            '_id': new_id(),
            'email': admin_email,
            'username': admin_username,
            'phone': admin_phone,
            'name': admin_name,
            'password_hash': hash_pw(admin_pw),
            'role': 'admin',
            'active': True,
            'created_at': iso(now_utc()),
        })
        logger.info(f'Seeded admin: {admin_email}')
    else:
        # keep password in sync with env
        if not verify_pw(admin_pw, existing.get('password_hash', '')):
            await db.users.update_one({'_id': existing['_id']}, {'$set': {'password_hash': hash_pw(admin_pw)}})


@app.on_event('startup')
async def on_startup():
    await bootstrap()


@app.middleware('http')
async def _lazy_bootstrap(request: Request, call_next):
    if not _bootstrapped:
        try:
            await bootstrap()
        except Exception as e:
            logger.warning(f'Bootstrap warning: {e}')
    return await call_next(request)


@app.on_event('shutdown')
async def shutdown_db_client():
    client.close()
