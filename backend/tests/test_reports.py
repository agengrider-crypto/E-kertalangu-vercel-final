"""Backend tests for the new Laporan (Reports) feature.

Covers:
- GET /api/reports/attendance with filters (date single/range, activity_id, q, status)
- Stats consistency vs. rows
- GET /api/reports/attendance/export/xlsx returns proper xlsx (2 sheets)
- RBAC: role 'peserta' must get 403
"""
import io
import os
import uuid
from datetime import date

import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL').rstrip('/')
API = f"{BASE_URL}/api"

ADMIN_EMAIL = 'agengpadma8@gmail.com'
ADMIN_PASS = 'jokam354'


@pytest.fixture(scope='module')
def admin_headers():
    r = requests.post(f'{API}/auth/login', json={'identifier': ADMIN_EMAIL, 'password': ADMIN_PASS})
    assert r.status_code == 200, r.text
    return {'Authorization': f"Bearer {r.json()['token']}", 'Content-Type': 'application/json'}


@pytest.fixture(scope='module')
def peserta_headers(admin_headers):
    """Create a fresh peserta user via /api/users and return its auth headers + id (for cleanup)."""
    uname = f'test_rpt_p_{uuid.uuid4().hex[:6]}'
    pw = 'Peserta123!'
    r = requests.post(f'{API}/users', headers=admin_headers, json={
        'name': 'TEST Peserta Rpt', 'username': uname, 'password': pw, 'role': 'peserta',
    })
    assert r.status_code == 200, r.text
    uid = r.json()['id']
    tok = requests.post(f'{API}/auth/login', json={'identifier': uname, 'password': pw}).json()['token']
    yield {'Authorization': f'Bearer {tok}', 'Content-Type': 'application/json'}
    requests.delete(f'{API}/users/{uid}', headers=admin_headers)


@pytest.fixture(scope='module')
def seeded(admin_headers):
    """Seed 1 activity + 3 participants + 3 attendance (hadir/izin/alpha). Cleanup after."""
    today = date.today().isoformat()
    aname = f'__QA_RPT__{uuid.uuid4().hex[:6]}'
    a = requests.post(f'{API}/activities', headers=admin_headers, json={
        'name': aname, 'type': 'pengajian_rutin', 'date': today,
        'start_time': '19:00', 'end_time': '20:30', 'location': 'QA'
    }).json()
    aid = a['id']

    pids = []
    for label in ['Andi', 'Budi', 'Citra']:
        p = requests.post(f'{API}/participants', headers=admin_headers, json={
            'name': f'__QA_RPT__{label}_{uuid.uuid4().hex[:4]}', 'gender': 'L',
            'duplicate_action': 'append'
        }).json()
        pids.append(p['id'])

    for pid, st in zip(pids, ['hadir', 'izin', 'alpha']):
        r = requests.post(f'{API}/attendance/manual', headers=admin_headers, json={
            'activity_id': aid, 'participant_id': pid, 'status': st
        })
        assert r.status_code == 200, r.text

    yield {'aid': aid, 'aname': aname, 'pids': pids, 'today': today}

    # cleanup
    requests.delete(f'{API}/activities/{aid}', headers=admin_headers)
    for pid in pids:
        requests.delete(f'{API}/participants/{pid}', headers=admin_headers)


# ----------------- Tests -----------------
def test_reports_requires_auth():
    r = requests.get(f'{API}/reports/attendance')
    assert r.status_code in (401, 403)


def test_reports_forbidden_for_peserta(peserta_headers):
    r = requests.get(f'{API}/reports/attendance', headers=peserta_headers)
    assert r.status_code == 403


def test_reports_xlsx_forbidden_for_peserta(peserta_headers):
    r = requests.get(f'{API}/reports/attendance/export/xlsx', headers=peserta_headers)
    assert r.status_code == 403


def test_reports_basic_all(admin_headers, seeded):
    r = requests.get(f'{API}/reports/attendance', headers=admin_headers,
                     params={'activity_id': seeded['aid']})
    assert r.status_code == 200
    d = r.json()
    assert 'rows' in d and 'stats' in d
    s = d['stats']
    # We seeded exactly 3 rows for this activity
    assert len(d['rows']) == 3
    assert s['hadir'] == 1 and s['izin'] == 1 and s['alpha'] == 1
    assert s['total'] == 3
    assert s['total_peserta'] == 3
    assert s['rate_hadir'] == round(1 / 3 * 100, 1)


def test_reports_date_single(admin_headers, seeded):
    r = requests.get(f'{API}/reports/attendance', headers=admin_headers,
                     params={'activity_id': seeded['aid'], 'date_from': seeded['today']})
    assert r.status_code == 200
    assert len(r.json()['rows']) == 3


def test_reports_date_range(admin_headers, seeded):
    r = requests.get(f'{API}/reports/attendance', headers=admin_headers, params={
        'activity_id': seeded['aid'],
        'date_from': seeded['today'], 'date_to': seeded['today'],
    })
    assert r.status_code == 200
    assert len(r.json()['rows']) == 3


def test_reports_status_filter(admin_headers, seeded):
    r = requests.get(f'{API}/reports/attendance', headers=admin_headers,
                     params={'activity_id': seeded['aid'], 'status': 'hadir'})
    assert r.status_code == 200
    d = r.json()
    assert len(d['rows']) == 1
    assert d['rows'][0]['status'] == 'hadir'
    assert d['stats']['hadir'] == 1
    assert d['stats']['izin'] == 0 and d['stats']['alpha'] == 0
    assert d['stats']['rate_hadir'] == 100.0


def test_reports_status_filter_alpha(admin_headers, seeded):
    r = requests.get(f'{API}/reports/attendance', headers=admin_headers,
                     params={'activity_id': seeded['aid'], 'status': 'alpha'})
    assert r.status_code == 200
    d = r.json()
    assert len(d['rows']) == 1 and d['rows'][0]['status'] == 'alpha'
    assert d['stats']['alpha'] == 1 and d['stats']['hadir'] == 0
    assert d['stats']['rate_hadir'] == 0


def test_reports_search_by_name(admin_headers, seeded):
    # search by unique tag we used
    r = requests.get(f'{API}/reports/attendance', headers=admin_headers,
                     params={'activity_id': seeded['aid'], 'q': '__QA_RPT__Andi'})
    assert r.status_code == 200
    d = r.json()
    assert len(d['rows']) == 1
    assert d['rows'][0]['participant_name'].startswith('__QA_RPT__Andi')


def test_reports_search_case_insensitive(admin_headers, seeded):
    r = requests.get(f'{API}/reports/attendance', headers=admin_headers,
                     params={'activity_id': seeded['aid'], 'q': '__qa_rpt__andi'})
    assert r.status_code == 200
    assert len(r.json()['rows']) == 1


def test_reports_empty_result(admin_headers, seeded):
    r = requests.get(f'{API}/reports/attendance', headers=admin_headers,
                     params={'activity_id': seeded['aid'], 'q': 'zzz_no_match_zzz'})
    assert r.status_code == 200
    d = r.json()
    assert d['rows'] == []
    assert d['stats']['total'] == 0
    assert d['stats']['total_peserta'] == 0
    assert d['stats']['rate_hadir'] == 0


def test_reports_xlsx(admin_headers, seeded):
    r = requests.get(f'{API}/reports/attendance/export/xlsx', headers=admin_headers,
                     params={'activity_id': seeded['aid']})
    assert r.status_code == 200
    ct = r.headers.get('content-type', '')
    assert 'spreadsheetml' in ct
    wb = load_workbook(io.BytesIO(r.content))
    assert 'Laporan Absensi' in wb.sheetnames
    assert 'Rekap' in wb.sheetnames
    ws = wb['Laporan Absensi']
    # header + 3 data rows
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == 'Tanggal Kegiatan'
    assert len(rows) == 4
    ws2 = wb['Rekap']
    rekap = {row[0]: row[1] for row in ws2.iter_rows(values_only=True)}
    assert rekap.get('Hadir') == 1
    assert rekap.get('Izin') == 1
    assert rekap.get('Alpha') == 1
    assert rekap.get('Total Peserta') == 3
